"""
generate_dashboard_data.py
──────────────────────────────────────────────────────────────────────────────
Reads the latest TIMB FCV daily summary file and injects live regional and
overview data into the Market Intelligence Dashboard 2026.html.

Updates two data blocks in the HTML:
  • /* RA_DATA_START */ … /* RA_DATA_END */   — Regional Competitiveness (P1)
  • /* OV_DATA_START */ … /* OV_DATA_END */   — Overview chart data (P0)

Run this script after each new daily FCV file arrives (or via run_2026_market_data.py).
"""

import os, sys, re, json, glob

import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass
from datetime import datetime, date, timedelta
from pathlib import Path

_WIN_ROOT = r"C:\Users\MikeBingo\Chevron Leaf Tobacco\Power BI Uploads - 2026 TIMB FCV Daily Files"
_LNX_ROOT = "/sessions/vibrant-tender-cray/mnt/Power BI Uploads - 2026 TIMB FCV Daily Files"
ROOT      = _WIN_ROOT if os.path.exists(_WIN_ROOT) else _LNX_ROOT
DAILY_DIR = ROOT   # daily FCV files live directly in root folder
DASHBOARD = os.path.join(ROOT, "Market Intelligence Dashboard 2026.html")

# ── Region configuration (static per season) ──────────────────────────────────
REGIONS = {
    'KAROI':   {'col': '#2D6A4F', 'crop25': 54001067, 'target': 1800000, 'sheet': 'KAROI'},
    'RUSAPE':  {'col': '#E67E22', 'crop25': 16091025, 'target': 800000,  'sheet': 'RUSAPE '},
    'BINDURA': {'col': '#3498DB', 'crop25': 7020698,  'target': 900000,  'sheet': 'BINDURA '},
    'HARARE':  {'col': '#9B59B6', 'crop25': 177040000, 'target': 300000,  'sheet': 'HARARE'},
}

STRIDE = 10   # columns per day block in regional sheets
CHEVRON_FRAG = 'Chevron Tobacco'


# ── Helpers ────────────────────────────────────────────────────────────────────
def day_num_from_path(p):
    m = re.search(r'day (\d+)', p, re.IGNORECASE)
    return int(m.group(1)) if m else 0

def find_latest_file():
    files = glob.glob(os.path.join(DAILY_DIR, "2026 daily fcv summary day *.xlsx"))
    if not files:
        raise RuntimeError("No daily FCV files found in root folder.")
    # Sort descending by day number; skip files that can't be opened (e.g. OneDrive cloud-only stubs)
    for f in sorted(files, key=day_num_from_path, reverse=True):
        try:
            import zipfile
            with open(f, 'rb') as fh:
                fh.read(4)   # just check we can read bytes
            return f, day_num_from_path(f)
        except OSError:
            continue
    raise RuntimeError("No readable daily FCV files found.")

def load_wb(path):
    import openpyxl
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise RuntimeError(f"Cannot read {path}: {e}")

def n(v, default=0.0):
    """Safe numeric conversion."""
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default

def wgt_price(mass, value):
    return value / mass if mass > 0 else 0.0


# ── Step 1: Build day → date mapping (scan all daily files) ───────────────────
def build_day_dates():
    files = glob.glob(os.path.join(DAILY_DIR, "2026 daily fcv summary day *.xlsx"))
    day_dates = {}
    for f in sorted(files, key=day_num_from_path):
        d = day_num_from_path(f)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb['Auction and Contract']
            for row in ws.iter_rows(values_only=True, min_row=1, max_row=6):
                for v in row:
                    if v and 'Date:' in str(v):
                        dt_str = str(v).replace('Date:', '').strip()
                        try:
                            day_dates[d] = datetime.strptime(dt_str, '%d/%m/%Y').date()
                        except ValueError:
                            pass
            wb.close()
        except Exception:
            pass
    return day_dates


# ── Step 2: Read one regional sheet ───────────────────────────────────────────
def read_regional_sheet(ws, max_day):
    """
    Returns:
      chev_daily: list of (mass, value) per day [0-indexed = day-1]
      mkt_daily:  list of (mass, value) per day
      peers:      dict company_name → (total_mass, total_value)
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find Chevron row and TOTAL row
    chev_row  = None
    total_row = None
    company_rows = []

    for r in rows:
        if not r[0]:
            continue
        name = str(r[0]).strip()
        if CHEVRON_FRAG in name:
            chev_row = r
        if name == 'TOTAL':
            total_row = r
        elif name not in ('Main Menu', 'Company') and not name.startswith('Previous') and name != 'TOTAL':
            company_rows.append(r)

    if chev_row is None or total_row is None:
        return [], [], {}

    chev_daily = []
    mkt_daily  = []

    for d in range(1, max_day + 1):
        c = 1 + (d - 1) * STRIDE
        cm = n(chev_row[c])
        cv = n(chev_row[c + 1])
        mm = n(total_row[c])
        mv = n(total_row[c + 1])
        chev_daily.append((cm, cv))
        mkt_daily.append((mm, mv))

    # Peers (seasonal total per company)
    peers = {}
    for r in company_rows:
        name = str(r[0]).strip()
        pm = sum(n(r[1 + (d - 1) * STRIDE]) for d in range(1, max_day + 1))
        pv = sum(n(r[2 + (d - 1) * STRIDE]) for d in range(1, max_day + 1))
        if pm > 0:
            peers[name] = (pm, pv)

    return chev_daily, mkt_daily, peers


# ── Step 3: Compute regional D object ─────────────────────────────────────────
def compute_region(region_key, cfg, chev_daily, mkt_daily, peers_raw,
                   day_dates, max_day):
    """Compute the full D[region] structure."""
    DAY = max_day   # latest day number

    # ── Determine period boundaries ──────────────────────────────────────────
    latest_date = day_dates.get(DAY)

    # WTD: from Monday of latest day's week
    if latest_date:
        monday = latest_date - timedelta(days=latest_date.weekday())
        wtd_start = next((d for d in range(1, DAY + 1) if day_dates.get(d, date(2000,1,1)) >= monday), 1)
    else:
        wtd_start = max(1, DAY - 4)

    # MTD: from 1st of latest day's month
    if latest_date:
        month_start = date(latest_date.year, latest_date.month, 1)
        mtd_start = next((d for d in range(1, DAY + 1) if day_dates.get(d, date(2000,1,1)) >= month_start), 1)
    else:
        mtd_start = max(1, DAY - 19)

    # March / April period boundaries for weekly breakdown
    # March = D1 to last day before April 1; April = from April 1 to DAY
    apr_start = None
    if day_dates:
        apr_start = next((d for d in range(1, DAY + 1)
                          if day_dates.get(d, date(2000,1,1)).month == 4), None)
    mar_end = (apr_start - 1) if apr_start and apr_start > 1 else DAY

    def sum_range(daily, start, end):
        """Sum (mass, value) for day-range [start, end] (1-indexed)."""
        tm = tv = 0.0
        for d in range(start, end + 1):
            m, v = daily[d - 1]
            tm += m; tv += v
        return tm, tv

    # ── Checkpoint arrays [Day, WTD, MTD, STD] ──────────────────────────────
    def checkpoints(daily):
        day_m, day_v = daily[DAY - 1]
        wtd_m, wtd_v = sum_range(daily, wtd_start, DAY)
        mtd_m, mtd_v = sum_range(daily, mtd_start, DAY)
        std_m, std_v = sum_range(daily, 1, DAY)
        return (
            [day_m, wtd_m, mtd_m, std_m],
            [day_v, wtd_v, mtd_v, std_v],
        )

    chev_mass, chev_val = checkpoints(chev_daily)
    mkt_mass,  mkt_val  = checkpoints(mkt_daily)

    def safe_prices(masses, values):
        return [round(wgt_price(m, v), 4) for m, v in zip(masses, values)]

    chev_price = safe_prices(chev_mass, chev_val)
    mkt_price  = safe_prices(mkt_mass,  mkt_val)
    price_gap  = [round(c - m, 4) if c and m else None
                  for c, m in zip(chev_price, mkt_price)]

    # ── STD summary ──────────────────────────────────────────────────────────
    std_chev_m, std_chev_v = chev_mass[3], chev_val[3]
    std_mkt_m,  std_mkt_v  = mkt_mass[3],  mkt_val[3]
    std_chev_p = wgt_price(std_chev_m, std_chev_v)
    std_mkt_p  = wgt_price(std_mkt_m,  std_mkt_v)
    price_gap_std = round(std_chev_p - std_mkt_p, 4)

    # ── Pace curves ──────────────────────────────────────────────────────────
    crop25 = cfg['crop25']
    target = cfg['target']

    # Cumulative masses per day
    chev_cum = []; mkt_cum = []
    cc = mc = 0.0
    for i in range(DAY):
        cc += chev_daily[i][0]
        mc += mkt_daily[i][0]
        chev_cum.append(cc)
        mkt_cum.append(mc)

    # Normalise: mkt vs crop25, chev vs target
    mkt_pace  = [round(mc / crop25 * 100, 3) for mc in mkt_cum]
    # Chevron pace: where chev had 0 that day show 0 not null
    chev_pace = [round(cc / target * 100, 3) if cc > 0 else 0.0 for cc in chev_cum]

    # Make final values exactly 100 if it reaches crop25/target
    day_labels = [f'D{d}' for d in range(1, DAY + 1)]

    # ── Pace gap ─────────────────────────────────────────────────────────────
    chev_pace_std = chev_cum[-1] / target * 100 if target else 0
    mkt_pace_std  = mkt_cum[-1]  / crop25 * 100 if crop25 else 0
    pace_gap_pp   = round(chev_pace_std - mkt_pace_std, 2)

    # ── Weekly breakdown ─────────────────────────────────────────────────────
    weekly = []
    if latest_date:
        # March period
        mar_cm, mar_cv = sum_range(chev_daily, 1, mar_end)
        mar_mm, mar_mv = sum_range(mkt_daily,  1, mar_end)
        mar_chev_p = wgt_price(mar_cm, mar_cv) or None
        mar_mkt_p  = wgt_price(mar_mm, mar_mv) or None
        mar_diff   = round(mar_chev_p - mar_mkt_p, 4) if mar_chev_p and mar_mkt_p else None
        mar_loss   = round(mar_cm * mar_diff, 0) if mar_diff and mar_cm else None
        weekly.append({'wk': '2026 Mar (D1-19)',
                       'mkt': mar_mkt_p, 'chev': mar_chev_p,
                       'mass': int(mar_cm), 'diff': mar_diff, 'loss': mar_loss})

        # April period
        if apr_start:
            apr_cm, apr_cv = sum_range(chev_daily, apr_start, DAY)
            apr_mm, apr_mv = sum_range(mkt_daily,  apr_start, DAY)
            apr_chev_p = wgt_price(apr_cm, apr_cv) or None
            apr_mkt_p  = wgt_price(apr_mm, apr_mv) or None
            apr_diff   = round(apr_chev_p - apr_mkt_p, 4) if apr_chev_p and apr_mkt_p else None
            apr_loss   = round(apr_cm * apr_diff, 0) if apr_diff and apr_cm else None
            weekly.append({'wk': f'2026 Apr (D{apr_start}-{DAY})',
                           'mkt': apr_mkt_p, 'chev': apr_chev_p,
                           'mass': int(apr_cm), 'diff': apr_diff, 'loss': apr_loss})

    season_loss = round(sum(w['loss'] for w in weekly if w['loss'] is not None), 0)

    # ── Risk tags ─────────────────────────────────────────────────────────────
    supply = '✓ ON TRACK' if pace_gap_pp >= 0 else '⚠ LAGGING'
    price_risk = '⚠ UNDERPAYING' if price_gap_std < -0.05 else '✓ COMPETITIVE'
    is_lagging = pace_gap_pp < 0
    is_under   = price_gap_std < -0.05
    if is_lagging and is_under:
        twin = '🔴 TWIN RISK'
    elif is_lagging or is_under:
        twin = '⚠ WATCH'
    else:
        twin = '✓ COMPETITIVE'

    # ── Peers list ────────────────────────────────────────────────────────────
    peers = []
    for name, (pm, pv) in sorted(peers_raw.items(), key=lambda x: -wgt_price(x[1][0], x[1][1])):
        peers.append({
            'n': name,
            'std': int(round(pm)),
            'price': round(wgt_price(pm, pv), 4),
            'me': CHEVRON_FRAG in name,
        })

    return {
        'col':          cfg['col'],
        'crop25':       crop25,
        'target':       target,
        'twin':         twin,
        'supply':       supply,
        'price_risk':   price_risk,
        'pace_gap_pp':  pace_gap_pp,
        'price_gap_std': price_gap_std,
        'mkt_mass':     [int(round(v)) for v in mkt_mass],
        'chev_mass':    [int(round(v)) for v in chev_mass],
        'mkt_price':    mkt_price,
        'chev_price':   chev_price,
        'price_gap':    price_gap,
        'mkt_pace_by_day':  mkt_pace,
        'chev_pace_by_day': chev_pace,
        'day_labels':   day_labels,
        'weekly':       weekly,
        'season_loss':  int(season_loss),
        'mkt_std':      int(round(std_mkt_m)),
        'mkt_price_std': round(std_mkt_p, 4),
        'peers':        peers,
    }


# ── Step 4: Build RA_DATA JS block ────────────────────────────────────────────
def build_ra_data(D_data, cr_data):
    """Produce JS: const D = {...}; const cr = [...]"""
    d_json = json.dumps(D_data, separators=(',', ':'))
    cr_json = json.dumps(cr_data, separators=(',', ':'))
    return f"const D={d_json};\nconst cr={cr_json};"


# ── Step 5: Build OV_DATA JS block ────────────────────────────────────────────
def build_ov_data(chev_daily_all, day_dates, max_day):
    """
    chev_daily_all: dict region→list of (mass, value) per day
    Produces: const cd=[...]; const DD={...};
    """
    # cd: daily total Chevron mass across all regions (only days with any volume)
    cd = []
    for d in range(1, max_day + 1):
        total = sum(chev_daily_all[r][d - 1][0] for r in chev_daily_all)
        if total > 0:
            dt = day_dates.get(d)
            cd.append({'d': d, 'm': int(round(total))})

    # DD: day → short date label (e.g. '29Apr')
    DD = {}
    for d, dt in sorted(day_dates.items()):
        if d <= max_day:
            DD[d] = f"{dt.day}{dt.strftime('%b')}"   # portable: no zero-pad on Windows or Linux

    cd_json = json.dumps(cd, separators=(',', ':'))
    dd_json = json.dumps(DD, separators=(',', ':'))
    return f"const cd={cd_json};\nconst DD={dd_json};"



CLEANED_2025 = os.path.join(ROOT, "Cleaned Tobacco Market Data 2025.xlsx")


def load_2025_comparison(max_day):
    """
    Read Cleaned Tobacco Market Data 2025.xlsx and return:
      reg25: dict region -> {'m': kg, 'v': usd, 'p': $/kg}  for D1..max_day
      auc25: dict floor  -> {'m': kg, 'v': usd, 'p': $/kg}  for D1..max_day
    Uses the same day-equivalent slice as the current 2026 season.
    """
    import openpyxl as _xl
    from collections import defaultdict as _dd

    def _n(v):
        try: return float(v) if v is not None else 0.0
        except: return 0.0

    try:
        wb25 = _xl.load_workbook(CLEANED_2025, read_only=True, data_only=True)
    except Exception as e:
        print("  WARNING: Cannot load 2025 data: " + str(e))
        return {}, {}

    # ── Contract by region ────────────────────────────────────────────────────
    reg25 = _dd(lambda: {'m': 0.0, 'v': 0.0})
    ws = wb25['Contractor Data 2025']
    for r in ws.iter_rows(values_only=True, min_row=2):
        region = str(r[0]).strip() if r[0] else None
        day    = int(r[2]) if r[2] else 0
        mass   = _n(r[4]); value = _n(r[5])
        if region and 1 <= day <= max_day:
            reg25[region]['m'] += mass
            reg25[region]['v'] += value

    reg25_out = {}
    for region, d in reg25.items():
        p = d['v'] / d['m'] if d['m'] > 0 else 0.0
        reg25_out[region] = {'m': d['m'], 'v': d['v'], 'p': p}

    # ── Auction by floor ──────────────────────────────────────────────────────
    # Cols: Auction Floor, Company, Day, % Mass, Ave.(US$/kg), Ave.BaleWeight, Bales Sold, Mass(kg), Value(US$)
    auc25 = _dd(lambda: {'m': 0.0, 'v': 0.0})
    ws2 = wb25['Auction Data 2025']
    for r in ws2.iter_rows(values_only=True, min_row=2):
        floor = str(r[0]).strip() if r[0] else None
        day   = int(r[2]) if r[2] else 0
        mass  = _n(r[7]); value = _n(r[8])
        if floor and 1 <= day <= max_day and mass > 0:
            auc25[floor]['m'] += mass
            auc25[floor]['v'] += value

    auc25_out = {}
    for floor, d in auc25.items():
        p = d['v'] / d['m'] if d['m'] > 0 else 0.0
        auc25_out[floor] = {'m': d['m'], 'v': d['v'], 'p': p}

    wb25.close()
    return reg25_out, auc25_out

# Extra sheets needed for daily totals (not in REGIONS regional analysis)
EXTRA_CONTRACT_SHEETS = {'MVURWI': 'MVURWI ', 'MARONDERA': 'MARONDERA'}
AUCTION_FLOOR_SHEETS  = ['TSF Daily Auction sales', 'ETF Daily Auction sales',
                          'PTSF Daily Auction sales']
FLOOR_STRIDE = 6


# ── Step 6: Extract P0 Season Overview data from xlsx ────────────────────────
def extract_p0_overview(wb, auc25_data=None):
    def nv(v):
        try: return float(v) if v is not None else 0.0
        except: return 0.0
    def safe_price(m, val):
        return val / m if m > 0 else 0.0

    out = {}

    # Auction and Contract sheet (per-floor 2026 + national 2025)
    ws = wb['Auction and Contract']
    rows = list(ws.iter_rows(values_only=True, min_row=5, max_row=10))
    mass_row = rows[1]; val_row = rows[2]
    tsf26_m = nv(mass_row[1]); tsf26_v = nv(val_row[1])
    etf26_m = nv(mass_row[2]); etf26_v = nv(val_row[2])
    ptf26_m = nv(mass_row[3]); ptf26_v = nv(val_row[3])
    auc26_m = nv(mass_row[4]); auc26_v = nv(val_row[4])
    con26_m = nv(mass_row[5]); con26_v = nv(val_row[5])
    nat26_m = nv(mass_row[6]); nat26_v = nv(val_row[6])
    nat25_m = nv(mass_row[7]); nat25_v = nv(val_row[7])

    out['tsf26_m'] = tsf26_m; out['tsf26_v'] = tsf26_v; out['tsf26_p'] = safe_price(tsf26_m, tsf26_v)
    out['etf26_m'] = etf26_m; out['etf26_v'] = etf26_v; out['etf26_p'] = safe_price(etf26_m, etf26_v)
    out['ptf26_m'] = ptf26_m; out['ptf26_v'] = ptf26_v; out['ptf26_p'] = safe_price(ptf26_m, ptf26_v)

    # 2025 auction: from load_2025_comparison if available, else Seasonal sheet
    if auc25_data:
        auc25_m = sum(d['m'] for d in auc25_data.values())
        auc25_v = sum(d['v'] for d in auc25_data.values())
    else:
        ws2 = wb['Seasonal Auction summary']
        rows2 = list(ws2.iter_rows(values_only=True, min_row=5, max_row=9))
        auc25_m = nv(rows2[1][2]); auc25_v = nv(rows2[2][2])

    out['auc26_m'] = auc26_m; out['auc26_p'] = safe_price(auc26_m, auc26_v)
    out['auc25_m'] = auc25_m; out['auc25_p'] = safe_price(auc25_m, auc25_v)

    # 2025 contract: sum from reg25_data (computed in main), else Seasonal sheet
    # (set after call via out['con25_m'] update in main)
    ws3 = wb['Seasonal Contract Sales']
    rows3 = list(ws3.iter_rows(values_only=True, min_row=5, max_row=9))
    con25_m = nv(rows3[1][2]); con25_v = nv(rows3[2][2])

    out['con26_m'] = con26_m; out['con26_p'] = safe_price(con26_m, con26_v)
    out['con25_m'] = con25_m; out['con25_p'] = safe_price(con25_m, con25_v)

    out['nat26_m'] = nat26_m; out['nat26_p'] = safe_price(nat26_m, nat26_v)
    out['nat25_m'] = nat25_m; out['nat25_p'] = safe_price(nat25_m, nat25_v)

    # SELLING POINT SUMMARY (regional 2026)
    ws = wb['SELLING POINT SUMMARY']
    rows4 = list(ws.iter_rows(values_only=True, min_row=5, max_row=20))
    EXCLUDE = {'GRAND TOTAL', 'Previous', 'Main Menu', 'Next', 'SELLING POINT', 'MUTOKO'}
    regions = []
    for r in rows4[1:]:
        if r[0] is None: continue
        nm = str(r[0]).strip()
        if nm in EXCLUDE or nm == '': continue
        co = int(nv(r[1])); m = nv(r[2]); v = nv(r[3]); p = nv(r[4])
        if m > 0:
            regions.append({'name': nm, 'co': co, 'm': m, 'v': v, 'p': p})
    out['regions'] = regions
    return out


# ── Step 7: Read extra contract sheet daily totals ────────────────────────────
def read_region_daily_totals(ws, max_day, stride=10):
    """Return list of (mass, value) per day using TOTAL row only."""
    rows = list(ws.iter_rows(values_only=True))
    total_row = None
    for r in rows:
        if r[0] and str(r[0]).strip() == 'TOTAL':
            total_row = r; break
    if total_row is None:
        return [(0.0, 0.0)] * max_day
    result = []
    for d in range(max_day):
        c = 1 + d * stride
        m = n(total_row[c]) if c < len(total_row) else 0.0
        v = n(total_row[c+1]) if c+1 < len(total_row) else 0.0
        result.append((m, v))
    return result


# ── Step 8: Read auction floor daily totals ───────────────────────────────────
def read_auction_floors_daily(wb, max_day):
    """Sum daily mass+value across all 3 floor sheets (STRIDE=6)."""
    totals = [(0.0, 0.0)] * max_day
    for sh in AUCTION_FLOOR_SHEETS:
        if sh not in wb.sheetnames: continue
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        total_row = next((r for r in rows if r[0] and 'TOTAL' in str(r[0]).upper()), None)
        if total_row is None: continue
        for d in range(max_day):
            c = 1 + d * FLOOR_STRIDE
            m = n(total_row[c]) if c < len(total_row) else 0.0
            v = n(total_row[c+1]) if c+1 < len(total_row) else 0.0
            totals[d] = (totals[d][0] + m, totals[d][1] + v)
    return totals


# ── Step 9: Build price chart SVG content ─────────────────────────────────────
def build_price_chart_content(daily_con, daily_auc, max_day):
    """
    daily_con: list of (mass, value) per day [all contract regions summed]
    daily_auc: list of (mass, value) per day [all auction floors summed]
    Returns full HTML block: <div class="price-line-wrap">...<svg>...</svg></div>
    """
    # Cumulative prices per day
    cum_cm = cum_cv = cum_am = cum_av = 0.0
    con_pts = []; auc_pts = []; nat_pts = []

    for d in range(max_day):
        cm, cv = daily_con[d]; am, av = daily_auc[d]
        cum_cm += cm; cum_cv += cv
        cum_am += am; cum_av += av
        cp = cum_cv / cum_cm if cum_cm > 0 else None
        ap = cum_av / cum_am if cum_am > 0 else None
        nm = cum_cm + cum_am; nv2 = cum_cv + cum_av
        np_ = nv2 / nm if nm > 0 else None
        con_pts.append(cp); auc_pts.append(ap); nat_pts.append(np_)

    # X scale: D1 at x=46, D_max at x=746
    W = 700.0
    def px(d_zero):  # d_zero is 0-indexed day
        return round(46.0 + d_zero * W / (max_day - 1), 1)

    # Y scale: price -> y (1.50 = 157.5, 3.50 = 27.0; 65.25px per dollar)
    Y_BOT = 157.5; Y_SCALE = 65.25
    def py(price):
        if price is None: return None
        y = Y_BOT - (price - 1.50) * Y_SCALE
        return round(max(10.0, min(164.0, y)), 1)

    # Determine day labels (every 5 days + last day)
    label_days = list(range(0, max_day, 5))  # 0, 5, 10, ...
    if (max_day - 1) not in label_days:
        label_days.append(max_day - 1)

    h = '  <!-- ====== DAILY PRICE TREND CHART ====== -->\n'
    h += '  <div class="price-line-wrap" style="border-color:#30363d;margin-bottom:18px">\n'
    h += '    <div class="plct"><div class="dl" style="background:var(--chv)"></div>'
    h += '2026 Cumulative Avg Price ($/kg) by Selling Day &nbsp;&middot;&nbsp; Contract &nbsp;&middot;&nbsp; Auction &nbsp;&middot;&nbsp; National</div>\n'
    h += '    <svg viewBox="0 0 760 200" xmlns="http://www.w3.org/2000/svg" style="width:100%;height:200px;display:block">\n'

    # Grid lines (horizontal — price levels)
    for price, label in [(1.50, '$1.50'), (2.00, '$2.00'), (2.50, '$2.50'),
                         (3.00, '$3.00'), (3.50, '$3.50')]:
        yg = py(price)
        h += '<line x1="46" y1="{}" x2="746" y2="{}" stroke="#21262d" stroke-width="1"/>\n'.format(yg, yg)
        h += '<text x="42" y="{}" text-anchor="end" fill="#7d8590" font-size="10">{}</text>\n'.format(round(yg + 3.5, 1), label)

    # Grid lines (vertical — day markers)
    for di in label_days:
        xv = px(di)
        lbl = 'D{}'.format(di + 1)
        h += '<line x1="{}" y1="14" x2="{}" y2="164" stroke="#21262d" stroke-width="1" stroke-dasharray="3,3"/>\n'.format(xv, xv)
        h += '<text x="{}" y="178" text-anchor="middle" fill="#7d8590" font-size="10">{}</text>\n'.format(xv, lbl)

    # X-axis baseline
    h += '<line x1="46" y1="164" x2="746" y2="164" stroke="#30363d" stroke-width="1"/>\n'

    # Helper: build polyline points string
    def polyline_pts(prices):
        pts = []
        for di, price in enumerate(prices):
            if price is None: continue
            pts.append('{},{}'.format(px(di), py(price)))
        return ' '.join(pts)

    # Contract line (green)
    h += '<polyline points="{}" fill="none" stroke="#2d9a4e" stroke-width="2.5" stroke-linejoin="round" stroke-linecap="round"/>\n'.format(polyline_pts(con_pts))
    # National line (gold)
    h += '<polyline points="{}" fill="none" stroke="#d4c060" stroke-width="2" stroke-linejoin="round" stroke-linecap="round"/>\n'.format(polyline_pts(nat_pts))
    # Auction line (blue)
    h += '<polyline points="{}" fill="none" stroke="#388bfd" stroke-width="2.5" stroke-linejoin="round" stroke-linecap="round"/>\n'.format(polyline_pts(auc_pts))

    # End-point dots and labels
    last = max_day - 1
    xlast = px(last)
    con_last = con_pts[last]; auc_last = auc_pts[last]; nat_last = nat_pts[last]
    if con_last:
        yc = py(con_last)
        h += '<circle cx="{}" cy="{}" r="4" fill="#2d9a4e"/>\n'.format(xlast, yc)
        h += '<text x="{}" y="{}" text-anchor="end" fill="#2d9a4e" font-size="9.5" font-weight="700">${:.2f}</text>\n'.format(xlast - 5, yc - 3, con_last)
    if nat_last:
        yn = py(nat_last)
        h += '<circle cx="{}" cy="{}" r="4" fill="#d4c060"/>\n'.format(xlast, yn)
        h += '<text x="{}" y="{}" text-anchor="end" fill="#d4c060" font-size="9.5" font-weight="700">${:.2f}</text>\n'.format(xlast - 5, yn + 11, nat_last)
    if auc_last:
        ya = py(auc_last)
        h += '<circle cx="{}" cy="{}" r="4" fill="#388bfd"/>\n'.format(xlast, ya)
        h += '<text x="{}" y="{}" text-anchor="end" fill="#388bfd" font-size="9.5" font-weight="700">${:.2f}</text>\n'.format(xlast - 5, ya - 3, auc_last)

    # Legend
    h += '<line x1="56" y1="189" x2="76" y2="189" stroke="#2d9a4e" stroke-width="2.5"/>\n'
    h += '<text x="81" y="193" fill="#7d8590" font-size="10">2026 Contract</text>\n'
    h += '<line x1="231" y1="189" x2="251" y2="189" stroke="#d4c060" stroke-width="2.5"/>\n'
    h += '<text x="256" y="193" fill="#7d8590" font-size="10">2026 National</text>\n'
    h += '<line x1="406" y1="189" x2="426" y2="189" stroke="#388bfd" stroke-width="2.5"/>\n'
    h += '<text x="431" y="193" fill="#7d8590" font-size="10">2026 Auction</text>\n'

    h += '    </svg>\n'
    h += '  </div>\n'
    return h


# ── Step 10: Build P0 header HTML ─────────────────────────────────────────────
def build_p0_hdr_html(day, end_date):
    ma = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
    ds = str(end_date.day).zfill(2) + ' ' + ma[end_date.month]
    return (
        '      <p>Chevron Tobacco &nbsp;&middot;&nbsp; Season Day 1&ndash;' + str(day) +
        ' &nbsp;&middot;&nbsp; 04 Mar &ndash; ' + ds + ' 2026</p>\n'
        '    </div>\n  </div>\n'
        '  <div style="display:flex;gap:9px;align-items:center">\n'
        '    <span class="badge">DAY ' + str(day) + '</span>\n'
        '    <span style="font-size:11px;color:var(--mut)">2026 vs 2025 Day 1&ndash;' + str(day) + ' equivalent</span>\n'
        '  </div>\n'
    )


# ── Step 11: Build P0 overview body HTML ──────────────────────────────────────
def build_p0_ov_html(ov, day, end_date, reg25=None, auc25=None):
    ma = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
    ds = str(end_date.day).zfill(2) + ' ' + ma[end_date.month]

    def fmt_m(kg): return '{:.1f}M'.format(kg / 1e6)
    def fmt_mn(kg): return '{:.2f}M'.format(kg / 1e6)
    def fmt_p(p): return '${:.2f}'.format(p)
    def fmt_v(usd): return '${:.1f}M'.format(usd / 1e6)

    def pct(new_v, old_v):
        if old_v == 0: return ('', 'up')
        c = (new_v - old_v) / old_v * 100
        if c >= 0: return ('&#9650; +{:.1f}%'.format(c), 'up')
        else: return ('&#9660; &minus;{:.1f}%'.format(abs(c)), 'dn')

    nat26m = ov['nat26_m']; nat25m = ov['nat25_m']
    nat26p = ov['nat26_p']; nat25p = ov['nat25_p']
    con26m = ov['con26_m']; con25m = ov['con25_m']
    con26p = ov['con26_p']; con25p = ov['con25_p']
    auc26m = ov['auc26_m']; auc25m = ov['auc25_m']
    auc26p = ov['auc26_p']; auc25p = ov['auc25_p']
    tsf26m = ov['tsf26_m']; tsf26v = ov['tsf26_v']; tsf26p = ov['tsf26_p']
    ptf26m = ov['ptf26_m']; ptf26v = ov['ptf26_v']; ptf26p = ov['ptf26_p']
    etf26m = ov['etf26_m']; etf26v = ov['etf26_v']; etf26p = ov['etf26_p']
    # 2025 floor data from load_2025_comparison (TSF=TSF, Premier=PTSF)
    if auc25:
        tsf_d  = auc25.get('TSF', {'m': 0, 'v': 0, 'p': 0})
        ptf_d  = auc25.get('Premier', {'m': 0, 'v': 0, 'p': 0})
        tsf25m = tsf_d['m']; tsf25v = tsf_d['v']; tsf25p = tsf_d['p']
        ptf25m = ptf_d['m']; ptf25v = ptf_d['v']; ptf25p = ptf_d['p']
    else:
        tsf25m = 3970000; tsf25v = 13748110.0; tsf25p = tsf25v / tsf25m
        ptf25m = 2520000; ptf25v = 8489880.0;  ptf25p = ptf25v / ptf25m

    nat_mc, nat_mc_cls = pct(nat26m, nat25m)
    nat_pc, nat_pc_cls = pct(nat26p, nat25p)
    con_mc, con_mc_cls = pct(con26m, con25m)
    con_pc, con_pc_cls = pct(con26p, con25p)
    auc_mc, auc_mc_cls = pct(auc26m, auc25m)
    auc_pc, auc_pc_cls = pct(auc26p, auc25p)
    tsf_mc, tsf_mc_cls = pct(tsf26m, tsf25m)
    tsf_pc, tsf_pc_cls = pct(tsf26p, tsf25p)
    ptf_mc, ptf_mc_cls = pct(ptf26m, ptf25m)
    ptf_pc, ptf_pc_cls = pct(ptf26p, ptf25p)

    RS = 'display:grid;grid-template-columns:52px 1fr 1fr 1fr'
    CM = 'padding:9px 0 9px 13px;font-size:11px;color:var(--mut);font-weight:600;display:flex;align-items:center'
    C5 = 'padding:9px 10px;text-align:right;font-size:12px;color:var(--mut)'
    CC = 'padding:9px 10px;text-align:right;font-size:12px;font-weight:700'

    h = ''
    h += '<div class="p0-topbar">\n'
    h += '  <div class="p0-topbar-title">TIMB FCV &mdash; 2026 SEASON OVERVIEW</div>\n'
    h += '  <div class="p0-topbar-sub">Season Day ' + str(day) + ' &nbsp;&middot;&nbsp; Full Season &nbsp;&middot;&nbsp; 04 Mar &ndash; ' + ds + ' 2026</div>\n'
    h += '</div>\n<div class="p0-body">\n\n'

    h += '  <!-- ====== OVERVIEW SUMMARY TABLE ====== -->\n'
    h += '  <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#888;margin-bottom:10px">'
    h += 'Overview &mdash; 2026 vs 2025 (Day 1&ndash;' + str(day) + ' equivalent)</div>\n'
    h += '  <div class="ovw-grid">\n\n'

    def ovw_card(cls, icon_cls, icon, title, m25, p25, m26, p26, color, mc, mc_cls, pc, pc_cls):
        s  = '    <div class="ovw-card ' + cls + '">\n'
        s += '      <div class="ovw-head ' + cls + '"><div class="ovw-icon ' + icon_cls + '">' + icon + '</div>' + title + '</div>\n'
        s += '      <div class="ovw-body">\n'
        s += '        <div style="' + RS + ';border-bottom:1px solid var(--brd)">\n'
        s += '          <div></div>\n'
        for lbl in ['2025', '2026', 'Change']:
            s += '          <div class="col-hdr" style="padding:5px 10px;font-size:10px;color:var(--mut);font-weight:600;text-align:right;display:flex;align-items:center;justify-content:flex-end">' + lbl + '</div>\n'
        s += '        </div>\n'
        s += '        <div style="' + RS + ';border-bottom:1px solid var(--brd)">\n'
        s += '          <div class="ovw-row" style="display:contents">'
        s += '<div style="' + CM + '">Mass</div>'
        s += '<div style="' + C5 + '">' + fmt_m(m25) + '</div>'
        s += '<div style="padding:9px 10px;text-align:right;font-size:13px;font-weight:700;color:' + color + '">' + fmt_m(m26) + '</div>'
        s += '<div style="' + CC + '" class="' + mc_cls + '">' + mc + '</div></div>\n'
        s += '        </div>\n'
        s += '        <div style="' + RS + '">\n'
        s += '          <div style="display:contents">'
        s += '<div style="' + CM + '">Price</div>'
        s += '<div style="' + C5 + '">' + fmt_p(p25) + '</div>'
        s += '<div style="padding:9px 10px;text-align:right;font-size:13px;font-weight:700;color:' + color + '">' + fmt_p(p26) + '</div>'
        s += '<div style="' + CC + '" class="' + pc_cls + '">' + pc + '</div></div>\n'
        s += '        </div>\n      </div>\n    </div>\n\n'
        return s

    h += ovw_card('nat', 'ni', '&Sigma;', 'National Crop',
                  nat25m, nat25p, nat26m, nat26p, '#d4c060',
                  nat_mc, nat_mc_cls, nat_pc, nat_pc_cls)
    h += ovw_card('con', 'ci', 'C', 'Contract',
                  con25m, con25p, con26m, con26p, 'var(--tel)',
                  con_mc, con_mc_cls, con_pc, con_pc_cls)
    h += ovw_card('auc', 'ai', 'A', 'Auction',
                  auc25m, auc25p, auc26m, auc26p, 'var(--blu)',
                  auc_mc, auc_mc_cls, auc_pc, auc_pc_cls)
    h += '  </div>\n\n'

    # Auction section
    auc_mvol, auc_mvol_cls = pct(auc26m, auc25m)
    auc_mprc, auc_mprc_cls = pct(auc26p, auc25p)
    h += '  <!-- ====== AUCTION SECTION ====== -->\n'
    h += '  <div class="sec-hdr">\n    <div class="sec-icon si-a">A</div>\n    <h2>AUCTION</h2>\n'
    h += '    <div class="sec-totals">\n'
    h += '      <span>2026: <strong>' + fmt_mn(auc26m) + ' kg</strong></span>\n'
    h += '      <span>2025: <strong>' + fmt_mn(auc25m) + ' kg</strong></span>\n'
    h += '      <span class="' + auc_mvol_cls + '" style="font-weight:700">' + auc_mvol + ' vol</span>\n'
    h += '      <span class="' + auc_mprc_cls + '" style="font-weight:700">' + auc_mprc + ' price</span>\n'
    h += '    </div>\n  </div>\n\n'

    def floor_card(name, note, m26, v26, p26, m25, v25, p25, is_new=False):
        mc, mc_cls = pct(m26, m25) if not is_new else ('', 'up')
        pc2, pc_cls2 = pct(p26, p25) if not is_new else ('', 'dn')
        s  = '    <div class="fc fl">\n      <div class="fn">' + name
        if note: s += ' <span style="font-weight:400;font-size:9px;opacity:.55">' + note + '</span>'
        s += '</div>\n      <div class="cmp">\n'
        s += '        <div><div class="yr">2026</div>'
        s += '<div class="fv y6b">' + fmt_mn(m26) + ' <span style="font-size:11px">kg</span></div>'
        s += '<div class="fp">' + fmt_v(v26) + ' &nbsp;&middot;&nbsp; ' + fmt_p(p26) + '/kg</div></div>\n'
        if is_new:
            s += '        <div><div class="yr">2025</div><div class="fv y5">&mdash;</div>'
            s += '<div class="fp" style="font-size:10px">No 2025 equivalent</div></div>\n'
        else:
            s += '        <div><div class="yr">2025</div>'
            s += '<div class="fv y5">' + fmt_mn(m25) + ' <span style="font-size:11px">kg</span></div>'
            s += '<div class="fp">' + fmt_v(v25) + ' &nbsp;&middot;&nbsp; ' + fmt_p(p25) + '/kg</div></div>\n'
        s += '      </div>\n      <div class="tags">'
        if is_new:
            s += '<span class="tag twn">New floor in 2026</span>'
        else:
            s += '<span class="tag t' + mc_cls + '">' + mc + ' vol</span>'
            s += '<span class="tag t' + pc_cls2 + '">' + pc2 + ' price</span>'
        s += '</div>\n    </div>\n'
        return s

    h += '  <div class="fc-grid">\n'
    h += floor_card('TSF Floor', None, tsf26m, tsf26v, tsf26p, tsf25m, tsf25v, tsf25p)
    h += floor_card('PTSF Floor', '= Premier 2025', ptf26m, ptf26v, ptf26p, ptf25m, ptf25v, ptf25p)
    h += floor_card('ETF Floor', 'New 2026', etf26m, etf26v, etf26p, 0, 0, 0, is_new=True)
    h += '  </div>\n\n'

    # Contract section
    con_mvol, con_mvol_cls = pct(con26m, con25m)
    con_mprc, con_mprc_cls = pct(con26p, con25p)
    h += '  <!-- ====== CONTRACT SECTION ====== -->\n'
    h += '  <div class="sec-hdr">\n    <div class="sec-icon si-c">C</div>\n    <h2>CONTRACT</h2>\n'
    h += '    <div class="sec-totals">\n'
    h += '      <span>2026: <strong>' + fmt_m(con26m) + ' kg</strong></span>\n'
    h += '      <span>2025: <strong>' + fmt_m(con25m) + ' kg</strong></span>\n'
    h += '      <span class="' + con_mvol_cls + '" style="font-weight:700">' + con_mvol + ' vol</span>\n'
    h += '      <span class="' + con_mprc_cls + '" style="font-weight:700">' + con_mprc + ' price</span>\n'
    h += '    </div>\n  </div>\n\n'

    h += '  <div class="fc-grid" style="grid-template-columns:repeat(3,1fr)">\n'
    for reg in ov['regions']:
        nm = reg['name']; co = reg['co']; m = reg['m']; v = reg['v']; p = reg['p']
        r25 = (reg25 or {}).get(nm, None)
        s  = '    <div class="fc rg">\n      <div class="fn">' + nm + '</div>\n'
        s += '      <div class="cmp">\n'
        s += '        <div><div class="yr">2026</div>'
        s += '<div class="fv y6g">' + fmt_m(m) + ' <span style="font-size:11px">kg</span></div>'
        s += '<div class="fp">' + fmt_v(v) + ' &nbsp;&middot;&nbsp; ' + fmt_p(p) + '/kg &nbsp;&middot;&nbsp; ' + str(co) + ' co</div></div>\n'
        if r25:
            r25m = r25['m']; r25p = r25['p']; r25v = r25m * r25p
            s += '        <div><div class="yr">2025</div>'
            s += '<div class="fv y5">' + fmt_m(r25m) + ' <span style="font-size:11px">kg</span></div>'
            s += '<div class="fp">' + fmt_v(r25v) + ' &nbsp;&middot;&nbsp; ' + fmt_p(r25p) + '/kg</div></div>\n'
            mc2, mc_cls2 = pct(m, r25m)
            pc2, pc_cls2 = pct(p, r25p)
            s += '      </div>\n      <div class="tags">'
            s += '<span class="tag t' + mc_cls2 + '">' + mc2 + '</span>'
            s += '<span class="tag t' + pc_cls2 + '">' + pc2 + ' price</span>'
        else:
            s += '      </div>\n      <div class="tags">'
            s += '<span class="tag twn">2026 Season-to-Date</span>'
        s += '</div>\n    </div>\n'
        h += s
    h += '  </div>\n'
    return h


# ── Step 12: Inject all blocks into HTML ──────────────────────────────────────
def _inject_html_block(html, start_marker, end_marker, content):
    s = html.find(start_marker)
    e = html.find(end_marker)
    if s == -1 or e == -1:
        print('  WARNING: marker not found: ' + start_marker)
        return html, False
    return html[:s + len(start_marker)] + '\n' + content + '\n' + html[e:], True


def _build_dashboard_html(ra_js, ov_js, p0_hdr_html, p0_ov_html, p0_chart_html):
    """Read the current dashboard HTML and apply all injections.

    Pure with respect to its inputs (no timestamps, no marker, no writes).
    Used both by the production publish path and the --verify idempotency check.
    """
    with open(DASHBOARD, 'r', encoding='utf-8') as f:
        html = f.read()

    for ms, me, new_js in [
        ('/* RA_DATA_START */', '/* RA_DATA_END */', ra_js),
        ('/* OV_DATA_START */', '/* OV_DATA_END */', ov_js),
    ]:
        if ms not in html:
            print('  WARNING: JS marker not found: ' + ms); continue
        start = html.index(ms) + len(ms)
        end   = html.index(me)
        html  = html[:start] + '\n' + new_js + '\n' + html[end:]

    html, _ = _inject_html_block(html, '<!-- P0_HDR_START -->',   '<!-- P0_HDR_END -->',   p0_hdr_html)
    html, _ = _inject_html_block(html, '<!-- P0_OV_START -->',    '<!-- P0_OV_END -->',    p0_ov_html)
    html, _ = _inject_html_block(html, '<!-- P0_CHART_START -->', '<!-- P0_CHART_END -->', p0_chart_html)
    return html


def _git_short_sha():
    """Return the short git SHA at ROOT, or 'nogit' if git is unavailable."""
    import subprocess as _sp
    try:
        out = _sp.check_output(
            ['git', 'rev-parse', '--short', 'HEAD'],
            cwd=ROOT, stderr=_sp.DEVNULL, timeout=5
        )
        return out.decode('utf-8', errors='replace').strip() or 'nogit'
    except Exception:
        return 'nogit'


def _apply_publish_marker(html):
    """Insert or replace <meta name="build-version" content="..."> in <head>.

    Returns (new_html, marker_value). The marker format is:
        <ISO-8601 UTC timestamp>|<git-short-sha>
    Live-deployment verification: GET the published URL and confirm the meta
    tag's content matches the marker_value printed by this run.
    """
    marker_value = '{0}|{1}'.format(
        datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        _git_short_sha()
    )
    meta_tag = '<meta name="build-version" content="{0}">'.format(marker_value)
    pattern = re.compile(r'<meta\s+name="build-version"\s+content="[^"]*"\s*/?>')
    if pattern.search(html):
        html = pattern.sub(meta_tag, html, count=1)
    else:
        # Insert immediately after the first <head ...> tag.
        head_re = re.compile(r'(<head\b[^>]*>)', re.IGNORECASE)
        if head_re.search(html):
            html = head_re.sub(lambda m: m.group(1) + '\n  ' + meta_tag, html, count=1)
        else:
            print('  WARNING: <head> not found; build marker not inserted.')
    return html, marker_value


def _write_dashboard_atomic(html, dest_path):
    """Atomic publish: write to <dest>.tmp in the same folder, then os.replace().

    Viewers (browsers, Excel previews, GitHub Desktop) never observe a
    partially-written dashboard. The replace step is retried on PermissionError
    to handle transient Windows file-lock contention from OneDrive sync,
    indexer, or a stale viewer.
    """
    import time as _time
    tmp_path = dest_path + '.tmp'
    # Write to tmp first. If this fails the dest is untouched.
    with open(tmp_path, 'w', encoding='utf-8') as f:
        f.write(html)
    # Atomic rename. On Windows this fails with PermissionError if a process
    # holds dest open without FILE_SHARE_DELETE — typically a browser/Excel
    # has the dashboard open. Retry, then surface a clear human-readable error.
    for attempt in range(1, 7):          # up to 6 attempts ≈ 75s of backoff
        try:
            os.replace(tmp_path, dest_path)
            return
        except PermissionError:
            if attempt == 6:
                # Clean up the orphan .tmp so OneDrive doesn't sync it as a sibling
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
                raise PermissionError(
                    "Dashboard file appears to be open or locked. "
                    "Close the browser/Excel preview and re-run."
                )
            print('  Dashboard locked (attempt {0}/6) - retrying in 15s...'.format(attempt))
            _time.sleep(15)


def inject_into_dashboard(ra_js, ov_js, p0_hdr_html, p0_ov_html, p0_chart_html):
    """Production publish path: build HTML, stamp marker, write atomically."""
    html = _build_dashboard_html(ra_js, ov_js, p0_hdr_html, p0_ov_html, p0_chart_html)
    html, marker_value = _apply_publish_marker(html)
    print('  Publish marker: ' + marker_value)
    _write_dashboard_atomic(html, DASHBOARD)


# ── Pipeline (data gathering, separated for reuse by --verify) ────────────────
def _compute_pipeline_outputs():
    """Run the full data pipeline against the latest daily FCV file.

    Returns (ra_js, ov_js, p0_hdr_html, p0_ov_html, p0_chart_html) — the five
    pieces consumed by _build_dashboard_html / inject_into_dashboard.

    This function is invoked once in production and twice in --verify mode.
    It MUST be deterministic given identical source data on disk.
    """
    latest_path, max_day = find_latest_file()
    print("\n  Latest file: Day " + str(max_day) + " -- " + os.path.basename(latest_path))

    day_dates = build_day_dates()
    print("  Day->date mapping: " + str(len(day_dates)) + " days")

    wb = load_wb(latest_path)

    # Regional analysis (P1)
    print("  Processing regional sheets...")
    D_data = {}; cr_data = []; chev_daily_all = {}

    for region_key, cfg in REGIONS.items():
        sheet_name = cfg['sheet']
        if sheet_name not in wb.sheetnames:
            print("  WARNING: Sheet '" + sheet_name + "' not found -- skipping " + region_key); continue
        ws = wb[sheet_name]
        chev_daily, mkt_daily, peers_raw = read_regional_sheet(ws, max_day)
        chev_daily_all[region_key] = chev_daily
        region_data = compute_region(region_key, cfg, chev_daily, mkt_daily, peers_raw, day_dates, max_day)
        D_data[region_key] = region_data
        chev_std = region_data['chev_mass'][3]
        mkt_std  = region_data['mkt_std']
        cr_data.append({'r': region_key, 'm': chev_std})
        print("    {:10s}: Chev={:>8,} kg  Mkt={:>12,} kg  PaceGap={:+.1f}pp".format(
            region_key, chev_std, mkt_std, region_data['pace_gap_pp']))

    ra_js = build_ra_data(D_data, cr_data)
    ov_js = build_ov_data(chev_daily_all, day_dates, max_day)

    # Contract daily totals from all 6 regions (for price chart)
    print("  Reading daily totals for price chart...")
    daily_con = [(0.0, 0.0)] * max_day
    all_con_sheets = dict(REGIONS)  # KAROI, RUSAPE, BINDURA, HARARE
    all_con_sheets.update({k: {'sheet': v} for k, v in EXTRA_CONTRACT_SHEETS.items()})
    for rk, rcfg in all_con_sheets.items():
        sh = rcfg['sheet']
        if sh not in wb.sheetnames: continue
        ws = wb[sh]
        daily_r = read_region_daily_totals(ws, max_day, stride=STRIDE)
        daily_con = [(daily_con[d][0] + daily_r[d][0],
                      daily_con[d][1] + daily_r[d][1]) for d in range(max_day)]
    con_total = sum(x[0] for x in daily_con)
    print("    Contract total: {:,.0f} kg".format(con_total))

    daily_auc = read_auction_floors_daily(wb, max_day)
    auc_total = sum(x[0] for x in daily_auc)
    print("    Auction total:  {:,.0f} kg".format(auc_total))

    # Load 2025 day-equivalent comparison data
    print("  Loading 2025 day-equivalent data (D1-" + str(max_day) + ")...")
    reg25, auc25 = load_2025_comparison(max_day)
    if reg25:
        for rk, rd in sorted(reg25.items()):
            print("    2025 {}: {:.1f}M kg @ ${:.3f}/kg".format(
                rk, rd['m']/1e6, rd['p']))
    # Recompute 2025 contract national total from reg25
    con25_m_real = sum(d['m'] for d in reg25.values())
    con25_v_real = sum(d['v'] for d in reg25.values())

    # P0 season overview data
    print("  Extracting P0 season overview data...")
    ov_data = extract_p0_overview(wb, auc25_data=auc25)
    # Override national 2025 contract with day-equivalent sum
    if con25_m_real > 0:
        ov_data['con25_m'] = con25_m_real
        ov_data['con25_p'] = con25_v_real / con25_m_real
        auc25_total_m = sum(d['m'] for d in auc25.values())
        auc25_total_v = sum(d['v'] for d in auc25.values())
        ov_data['nat25_m'] = con25_m_real + auc25_total_m
        ov_data['nat25_p'] = (con25_v_real + auc25_total_v) / ov_data['nat25_m'] if ov_data['nat25_m'] > 0 else 0
    print("  National 2026: {:.1f}M kg @ ${:.3f}/kg".format(
        ov_data['nat26_m']/1e6, ov_data['nat26_p']))
    print("  National 2025: {:.1f}M kg @ ${:.3f}/kg".format(
        ov_data['nat25_m']/1e6, ov_data['nat25_p']))

    # Aggregate season totals → exposed as JS OV_AGG (appended to the OV_DATA
    # block). Lets the dashboard's static rendering code (sidebar, price chart)
    # read today's numbers instead of stale hardcoded values.
    ov_agg = {
        'nat26_m': ov_data['nat26_m'], 'nat26_p': ov_data['nat26_p'],
        'auc26_m': ov_data['auc26_m'], 'auc26_p': ov_data['auc26_p'],
        'con26_m': ov_data['con26_m'], 'con26_p': ov_data['con26_p'],
        'nat25_m': ov_data['nat25_m'], 'nat25_p': ov_data['nat25_p'],
        'auc25_m': ov_data['auc25_m'], 'auc25_p': ov_data['auc25_p'],
        'con25_m': ov_data['con25_m'], 'con25_p': ov_data['con25_p'],
    }
    ov_js = ov_js + "\nconst OV_AGG=" + json.dumps(ov_agg, separators=(',', ':')) + ";"

    end_date = day_dates.get(max_day)
    if end_date is None:
        from datetime import date as _date; end_date = _date.today()

    p0_hdr_html   = build_p0_hdr_html(max_day, end_date)
    p0_ov_html    = build_p0_ov_html(ov_data, max_day, end_date, reg25=reg25, auc25=auc25)
    p0_chart_html = build_price_chart_content(daily_con, daily_auc, max_day)

    return ra_js, ov_js, p0_hdr_html, p0_ov_html, p0_chart_html


# ── --verify: idempotency check ──────────────────────────────────────────────
def _run_verify():
    """Run the pipeline twice, build HTML twice (no marker, no write), and
    require the two outputs to be byte-identical. Exits 0 on match, 1 on diff.
    """
    print("** VERIFY MODE: running pipeline twice for idempotency check **\n")
    print("[Pass 1]")
    pieces1 = _compute_pipeline_outputs()
    html1   = _build_dashboard_html(*pieces1)
    print("\n[Pass 2]")
    pieces2 = _compute_pipeline_outputs()
    html2   = _build_dashboard_html(*pieces2)
    if html1 == html2:
        print("\n  + Idempotency verified: both passes produced identical HTML "
              "({0:,} bytes).".format(len(html1)))
        sys.exit(0)
    # Save both for human inspection
    import tempfile
    tdir = tempfile.gettempdir()
    p1 = os.path.join(tdir, 'verify_pass1.html')
    p2 = os.path.join(tdir, 'verify_pass2.html')
    with open(p1, 'w', encoding='utf-8') as f: f.write(html1)
    with open(p2, 'w', encoding='utf-8') as f: f.write(html2)
    print("\n  ! Idempotency FAILED: pass 1 ({0:,} bytes) != pass 2 ({1:,} bytes).".format(
        len(html1), len(html2)))
    print('    Pass 1 saved: ' + p1)
    print('    Pass 2 saved: ' + p2)
    print('    Diff with: fc /N "{0}" "{1}"   (Windows)'.format(p1, p2))
    print('         or:   diff "{0}" "{1}"   (Linux)'.format(p1, p2))
    sys.exit(1)


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("generate_dashboard_data.py -- Regional & Overview Live Data")
    print("=" * 60)

    if '--verify' in sys.argv:
        _run_verify()
        return

    pieces = _compute_pipeline_outputs()
    print("  Injecting into dashboard HTML...")
    inject_into_dashboard(*pieces)
    print("\n  + Dashboard updated successfully.")
    print("    Dashboard: " + DASHBOARD)
    print("=" * 60)


if __name__ == '__main__':
    main()
