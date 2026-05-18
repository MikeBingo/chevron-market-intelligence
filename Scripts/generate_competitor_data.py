"""
generate_competitor_data.py
Data sources:
  CONTRACT -> Regional sheets (KAROI, RUSAPE, BINDURA, HARARE)  STRIDE=10
  AUCTION  -> TSF + ETF + PTSF daily auction sheets             STRIDE=6
No SS/COM sector split. Big established merchants excluded.
"""

import os, sys, re, json, glob, argparse

import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass
from datetime import datetime
import openpyxl

_WIN_ROOT = r"C:\Users\MikeBingo\Chevron Leaf Tobacco\Power BI Uploads - 2026 TIMB FCV Daily Files"
_LNX_ROOT = "/sessions/vibrant-tender-cray/mnt/Power BI Uploads - 2026 TIMB FCV Daily Files"
ROOT      = _WIN_ROOT if os.path.exists(_WIN_ROOT) else _LNX_ROOT
DAILY_DIR = ROOT
DASHBOARD = os.path.join(ROOT, "Market Intelligence Dashboard 2026.html")

REGION_SHEETS  = ['KAROI', 'RUSAPE ', 'BINDURA ', 'HARARE', 'MVURWI ', 'MARONDERA', 'MUTOKO']
CON_STRIDE     = 10
AUCTION_SHEETS = ['TSF Daily Auction sales', 'ETF Daily Auction sales', 'PTSF Daily Auction sales']
AUC_STRIDE     = 6

CONTRACT_MAP = {
    'Inter  Continental Leaf': ('ilt',    'ILT (Core)'),
    'Intercontinental Leaf':   ('ilt',    'ILT (Core)'),
    'Hurudza Leaf':            ('ilt',    'Hurudza'),
    'Vision Leaf':             ('ilt',    'Vision Leaf'),
    'Servemox':                ('ilt',    'Servemox'),
    'Aqua Tobacco':            ('ilt',    'Aqua'),
    'Swan Valley Tobacco':     ('swan',   'Swan Valley'),
    'Sub-Sahara Tobacco P/L':  ('swan',   'Sub-Sahara'),
    'Ethical Leaf Tobacco':    ('eth',    'Ethical Leaf'),
    'Tobacco Corporation':     ('eth',    'Tobacco Corp'),
    'Global Leaf':             ('global', 'Global Leaf'),
    'Agro leaf Tobacco':       ('shasha', 'Agro Leaf'),
    'Chevron Tobacco Company': ('chev',   'Chevron'),
}

AUCTION_MAP = {
    'Intercontinental Leaf Tobacco': ('ilt',    'ILT (Core)'),
    'Country Agro':                  ('ilt',    'Country Agro'),
    'Ethical Leaf Tobacco':          ('eth',    'Ethical Leaf'),
    'Tobacco Corporation':           ('eth',    'Tobacco Corp'),
    'Sub-Sahara':                    ('swan',   'Sub-Sahara'),
    'Global Leaf Tobacco':           ('global', 'Global Leaf'),
    'Dobre Tobacco':                 ('dobre',  'Dobre'),
    'Ratenta Enterprises':           ('dobre',  'Rantenta'),
    'Chevron Tobacco P/L':           ('chev',   'Chevron'),
}

GROUP_META = {
    'ilt':    {'name': 'ILT Group',     'label': 'Dominant Threat',  'labelCol': '#9B1C1C', 'col': '#1A3560', 'rank': 1},
    'eth':    {'name': 'Ethical Group', 'label': 'Auction Dominant', 'labelCol': '#5B21B6', 'col': '#6B3FA0', 'rank': 2},
    'swan':   {'name': 'Swan Valley',   'label': 'Contract Only',    'labelCol': '#065F46', 'col': '#2D6A4F', 'rank': 3},
    'global': {'name': 'Global Leaf',   'label': 'Steady Contract',  'labelCol': '#0F766E', 'col': '#16A085', 'rank': 4},
    'dobre':  {'name': 'Dobre Group',   'label': 'Auction Only',     'labelCol': '#B91C1C', 'col': '#C0392B', 'rank': 5},
    'shasha': {'name': 'Shasha Group',  'label': 'Contract Signal',  'labelCol': '#B45309', 'col': '#E67E22', 'rank': 6},
    'chev':   {'name': 'Chevron',       'label': 'Baseline',         'labelCol': '#B8860B', 'col': '#B8860B', 'rank': 0},
}

def find_latest_file():
    files = glob.glob(os.path.join(DAILY_DIR, "2026 daily fcv summary day *.xlsx"))
    if not files:
        sys.exit("No daily FCV files found.")
    def day_num(p):
        m = re.search(r'day (\d+)', p, re.IGNORECASE)
        return int(m.group(1)) if m else 0
    for f in sorted(files, key=day_num, reverse=True):
        try:
            with open(f, 'rb') as fh:
                fh.read(4)
            return f, day_num(f)
        except OSError:
            continue
    sys.exit("No readable daily FCV files found.")

def load_wb(path):
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        sys.exit(f"Cannot read {path}: {e}")

def n(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0

def match_con(name):
    cn = str(name).strip()
    for frag, (grp, disp) in CONTRACT_MAP.items():
        if frag.lower() in cn.lower():
            return grp, disp
    return None, None

def match_auc(name):
    cn = str(name).strip()
    for frag, (grp, disp) in AUCTION_MAP.items():
        if frag.lower() in cn.lower():
            return grp, disp
    return None, None

def wgt_avg(mass, value):
    return value / mass if mass > 0 else 0.0

def new_group():
    return {'con_m': 0.0, 'con_v': 0.0, 'auc_m': 0.0, 'auc_v': 0.0, 'members': {}}

def add_con(groups, grp, disp, m, v):
    g = groups.setdefault(grp, new_group())
    g['con_m'] += m; g['con_v'] += v
    mem = g['members'].setdefault(disp, {'con_m': 0.0, 'con_v': 0.0, 'auc_m': 0.0, 'auc_v': 0.0})
    mem['con_m'] += m; mem['con_v'] += v

def add_auc(groups, grp, disp, m, v):
    g = groups.setdefault(grp, new_group())
    g['auc_m'] += m; g['auc_v'] += v
    mem = g['members'].setdefault(disp, {'con_m': 0.0, 'con_v': 0.0, 'auc_m': 0.0, 'auc_v': 0.0})
    mem['auc_m'] += m; mem['auc_v'] += v

def extract(path, day):
    wb = load_wb(path)
    groups = {}

    # 1. Contract from regional sheets
    print("  Reading contract from regional sheets...")
    for sheet_name in REGION_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"    WARNING: sheet '{sheet_name.strip()}' not found")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        captured = 0
        for r in rows:
            if not r[0]:
                continue
            name = str(r[0]).strip()
            if name in ('Main Menu', 'Company', 'TOTAL', 'Previous') or name.startswith('Previous'):
                continue
            grp, disp = match_con(name)
            if not grp:
                continue
            m = sum(n(r[1 + (d-1)*CON_STRIDE]) for d in range(1, day+1) if len(r) > 1+(d-1)*CON_STRIDE)
            v = sum(n(r[2 + (d-1)*CON_STRIDE]) for d in range(1, day+1) if len(r) > 2+(d-1)*CON_STRIDE)
            if m > 0:
                add_con(groups, grp, disp, m, v)
                captured += m
        print(f"    {sheet_name.strip():10s}: {captured:>12,.0f} kg")

    # 2. Auction from TSF + ETF + PTSF
    print("  Reading auction from floor sheets...")
    for sheet_name in AUCTION_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"    WARNING: sheet '{sheet_name}' not found")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        captured = 0
        for r in rows:
            if not r[0]:
                continue
            name = str(r[0]).strip()
            if name in ('Main Menu', 'Company', 'TOTAL') or not name:
                continue
            grp, disp = match_auc(name)
            if not grp:
                continue
            m = sum(n(r[1 + (d-1)*AUC_STRIDE]) for d in range(1, day+1) if len(r) > 1+(d-1)*AUC_STRIDE)
            v = sum(n(r[2 + (d-1)*AUC_STRIDE]) for d in range(1, day+1) if len(r) > 2+(d-1)*AUC_STRIDE)
            if m > 0:
                add_auc(groups, grp, disp, m, v)
                captured += m
        print(f"    {sheet_name:35s}: {captured:>10,.0f} kg")

    wb.close()

    # 3. Separate Chevron
    chev_raw = groups.pop('chev', new_group())
    chev_m   = chev_raw['con_m'] + chev_raw['auc_m']
    chev_v   = chev_raw['con_v'] + chev_raw['auc_v']

    # 4. Build group output
    out_groups = []
    for grp_key, meta in sorted(GROUP_META.items(), key=lambda x: x[1]['rank']):
        if grp_key == 'chev':
            continue
        g = groups.get(grp_key, new_group())
        tot_m = g['con_m'] + g['auc_m']
        tot_v = g['con_v'] + g['auc_v']
        members = []
        for mn, md in g['members'].items():
            members.append({
                'co':    mn,
                'con_m': round(md['con_m']),
                'con_p': round(wgt_avg(md['con_m'], md['con_v']), 4),
                'auc_m': round(md['auc_m']),
                'auc_p': round(wgt_avg(md['auc_m'], md['auc_v']), 4),
            })
        members.sort(key=lambda x: -(x['con_m'] + x['auc_m']))
        out_groups.append({
            'name':     meta['name'],
            'label':    meta['label'],
            'labelCol': meta['labelCol'],
            'col':      meta['col'],
            'rank':     meta['rank'],
            'm':        round(tot_m),
            'p':        round(wgt_avg(tot_m, tot_v), 4),
            'con_m':    round(g['con_m']),
            'con_p':    round(wgt_avg(g['con_m'], g['con_v']), 4),
            'auc_m':    round(g['auc_m']),
            'auc_p':    round(wgt_avg(g['auc_m'], g['auc_v']), 4),
            'members':  members,
        })
    out_groups.sort(key=lambda x: -x['m'])

    # 5. Channel totals
    comp_con_m = sum(g['con_m'] for g in out_groups)
    comp_con_v = sum(g['con_m'] * g['con_p'] for g in out_groups if g['con_p'])
    comp_auc_m = sum(g['auc_m'] for g in out_groups)
    comp_auc_v = sum(g['auc_m'] * g['auc_p'] for g in out_groups if g['auc_p'])
    comp_m     = sum(g['m']     for g in out_groups)
    comp_v     = sum(g['m']     * g['p'] for g in out_groups if g['p'])

    # 6. Date from file
    date_str = '--'
    try:
        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if 'Auction and Contract' in wb2.sheetnames:
            for r in wb2['Auction and Contract'].iter_rows(values_only=True, max_row=10):
                for cell in r:
                    if cell and 'Date:' in str(cell):
                        date_str = str(cell).replace('Date:', '').strip()
                        break
        wb2.close()
    except Exception:
        pass

    return {
        'day':       day,
        'date':      date_str,
        'generated': datetime.now().strftime('%d %b %Y %H:%M'),
        'chev': {
            'm':     round(chev_m),
            'p':     round(wgt_avg(chev_m, chev_v), 4),
            'con_m': round(chev_raw['con_m']),
            'con_p': round(wgt_avg(chev_raw['con_m'], chev_raw['con_v']), 4),
            'auc_m': round(chev_raw['auc_m']),
            'auc_p': round(wgt_avg(chev_raw['auc_m'], chev_raw['auc_v']), 4),
        },
        'chan': {
            'contract': {
                'comp_m': round(comp_con_m),
                'comp_p': round(wgt_avg(comp_con_m, comp_con_v), 4),
                'label':  str(sum(1 for g in out_groups if g['con_m'] > 0)) + ' groups active',
            },
            'auction': {
                'comp_m': round(comp_auc_m),
                'comp_p': round(wgt_avg(comp_auc_m, comp_auc_v), 4),
                'label':  str(sum(1 for g in out_groups if g['auc_m'] > 0)) + ' groups active',
            },
        },
        'comp_total': {
            'm': round(comp_m),
            'p': round(comp_v / comp_m, 4) if comp_m else 0,
        },
        'groups': out_groups,
    }

CLEANED_2025 = os.path.join(ROOT, "Cleaned Tobacco Market Data 2025.xlsx")

def load_chev_2025(max_day):
    """Load Chevron 2025 contract + auction figures up to max_day from cleaned 2025 data."""
    ref = {'con_m': 0, 'con_v': 0, 'auc_m': 0, 'auc_v': 0}
    try:
        wb25 = openpyxl.load_workbook(CLEANED_2025, read_only=True, data_only=True)
        # Contract: Region[0], Company[1], Day[2], Date[3], Mass[4], Value[5]
        if 'Contractor Data 2025' in wb25.sheetnames:
            ws = wb25['Contractor Data 2025']
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                company = str(row[1] or '').strip()
                if 'chevron' not in company.lower():
                    continue
                day = row[2]
                if day is None:
                    continue
                try:
                    day = int(day)
                except (ValueError, TypeError):
                    continue
                if day < 1 or day > max_day:
                    continue
                try:
                    ref['con_m'] += float(row[4] or 0)
                    ref['con_v'] += float(row[5] or 0)
                except (TypeError, ValueError):
                    pass
        # Auction: Floor[0], Company[1], Day[2], %[3], Price[4], BaleWt[5], Bales[6], Mass[7], Value[8]
        if 'Auction Data 2025' in wb25.sheetnames:
            ws = wb25['Auction Data 2025']
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                company = str(row[1] or '').strip()
                if 'chevron' not in company.lower():
                    continue
                day = row[2]
                if day is None:
                    continue
                try:
                    day = int(day)
                except (ValueError, TypeError):
                    continue
                if day < 1 or day > max_day:
                    continue
                try:
                    ref['auc_m'] += float(row[7] or 0)
                    ref['auc_v'] += float(row[8] or 0)
                except (TypeError, ValueError):
                    pass
        wb25.close()
    except Exception as e:
        print('Warning: could not load Chevron 2025 data: {}'.format(e))
        # Fall back to D1-39 hardcoded values
        return {'con_m': 229732, 'con_p': 3.188, 'auc_m': 160503, 'auc_p': 3.137,
                'tot_m': 390235, 'tot_p': 3.167}
    con_m = ref['con_m']
    con_p = ref['con_v'] / con_m if con_m else 0
    auc_m = ref['auc_m']
    auc_p = ref['auc_v'] / auc_m if auc_m else 0
    tot_m = con_m + auc_m
    tot_v = ref['con_v'] + ref['auc_v']
    tot_p = tot_v / tot_m if tot_m else 0
    return {
        'con_m': round(con_m), 'con_p': round(con_p, 3),
        'auc_m': round(auc_m), 'auc_p': round(auc_p, 3),
        'tot_m': round(tot_m), 'tot_p': round(tot_p, 3),
    }

def _pct(n, d):
    return (n - d) / d * 100 if d else 0.0

def _chg_html(v):
    arrow = 'up' if v >= 0 else 'dn'
    sign  = '+' if v >= 0 else chr(8722)   # minus sign U+2212
    tri   = '▲ ' if v >= 0 else '▼ −'
    if v >= 0:
        return '<div class="vv up" style="font-size:12px">▲ +' + '{:.1f}'.format(abs(v)) + '%</div>'
    else:
        return '<div class="vv dn" style="font-size:12px">▼ −' + '{:.1f}'.format(abs(v)) + '%</div>'

def _month_abbr(date_str):
    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    try:
        d, m, y = date_str.split('/')
        mo = months[int(m) - 1]
        return '{} {} {}'.format(d, mo, y), '{} {} {}'.format(d, mo.upper(), y)
    except Exception:
        return date_str, date_str.upper()

def _row_html(m25, p25, m26, p26, m_chg, p_chg, vv_cls):
    lines = []
    lines.append('      <div class="cb">')
    lines.append('        <div class="cr">')
    lines.append('          <div class="cl">Mass</div>')
    lines.append('          <div class="cv"><div class="vy">{:,.0f} kg</div></div>'.format(m25))
    lines.append('          <div class="cv"><div class="vv {}">{:,.0f}</div></div>'.format(vv_cls, m26))
    lines.append('          <div class="cv">{}</div>'.format(_chg_html(m_chg)))
    lines.append('        </div>')
    lines.append('        <div class="cr">')
    lines.append('          <div class="cl">Price</div>')
    lines.append('          <div class="cv"><div class="vy">${:.2f}/kg</div></div>'.format(p25))
    lines.append('          <div class="cv"><div class="vv {}">${:.2f}</div></div>'.format(vv_cls, p26))
    lines.append('          <div class="cv">{}</div>'.format(_chg_html(p_chg)))
    lines.append('        </div>')
    lines.append('      </div>')
    return '\n'.join(lines)

def _card_html(letter, label, bg_col, txt_col, m25, p25, m26, p26, m_chg, p_chg, vv_cls):
    icon = ('<span style="width:18px;height:18px;border-radius:4px;'
            'background:{};color:{};font-size:9px;font-weight:900;'
            'display:flex;align-items:center;justify-content:center">'
            '{}</span>').format(bg_col, txt_col, letter)
    lines = []
    lines.append('    <div class="cc {}">'.format(vv_cls.replace('v','').strip('gba') or 'con'))
    lines.append('      <div class="ch">{}{}</div>'.format(icon, label))
    lines.append('      <div class="cch">')
    lines.append('        <div class="chc"></div>')
    lines.append('        <div class="chc">2025</div><div class="chc">2026</div><div class="chc">Change</div>')
    lines.append('      </div>')
    lines.append(_row_html(m25, p25, m26, p26, m_chg, p_chg, vv_cls))
    lines.append('    </div>')
    return '\n'.join(lines)

def build_chev_context_html(chev, day, date_str, ref=None):
    r = ref if ref is not None else {'con_m': 229732, 'con_p': 3.188,
                                      'auc_m': 160503, 'auc_p': 3.137,
                                      'tot_m': 390235, 'tot_p': 3.167}
    con_m_chg = _pct(chev['con_m'], r['con_m'])
    con_p_chg = _pct(chev['con_p'], r['con_p'])
    auc_m_chg = _pct(chev['auc_m'], r['auc_m'])
    auc_p_chg = _pct(chev['auc_p'], r['auc_p'])
    tot_m_chg = _pct(chev['m'],     r['tot_m'])
    tot_p_chg = _pct(chev['p'],     r['tot_p'])

    tot_disp  = '{:.2f}M kg'.format(chev['m'] / 1e6)
    vol_cls   = 'up' if tot_m_chg >= 0 else 'dn'
    pr_cls    = 'up' if tot_p_chg >= 0 else 'dn'
    vol_arrow = ('▲ +' if tot_m_chg >= 0 else '▼ −') + '{:.1f}%'.format(abs(tot_m_chg))
    pr_arrow  = ('▲ +' if tot_p_chg >= 0 else '▼ −') + '{:.1f}%'.format(abs(tot_p_chg))

    con_card = _card_html('C', 'Contract', '#0d2918', 'var(--grn)',
                          r['con_m'], r['con_p'], chev['con_m'], chev['con_p'],
                          con_m_chg, con_p_chg, 'gv')
    auc_card = _card_html('A', 'Auction',  '#0d1f3c', 'var(--blu)',
                          r['auc_m'], r['auc_p'], chev['auc_m'], chev['auc_p'],
                          auc_m_chg, auc_p_chg, 'bv')
    tot_card = _card_html('Σ', 'Total', '#1a1400', 'var(--chv)',
                          r['tot_m'], r['tot_p'], chev['m'], chev['p'],
                          tot_m_chg, tot_p_chg, 'av')

    lines = []
    lines.append('  <!-- ====== CHEVRON CONTEXT ====== -->')
    lines.append('  <div class="sec-hdr" style="margin-top:22px">')
    lines.append('    <div class="sec-icon si-x">CT</div>')
    lines.append('    <h2>CHEVRON CONTEXT</h2>')
    lines.append('    <div class="sec-totals">')
    lines.append('      <span>2026 total: <strong>{}</strong></span>'.format(tot_disp))
    lines.append('      <span>2025 total: <strong>{}</strong> (Day 1–{})</span>'.format(
          '{:.0f}K kg'.format(r['tot_m'] / 1e3), day))
    lines.append('      <span class="{}" style="font-weight:700">{} vol &nbsp;</span>'
                 '<span class="{}" style="font-weight:700">{} price</span>'.format(
                     vol_cls, vol_arrow, pr_cls, pr_arrow))
    lines.append('    </div>')
    lines.append('  </div>')
    lines.append('')
    lines.append('  <div class="chv-grid">')
    lines.append(con_card)
    lines.append('')
    lines.append(auc_card)
    lines.append('')
    lines.append(tot_card)
    lines.append('  </div>')
    lines.append('')
    lines.append('</div><!-- /p0-body -->')
    return '\n'.join(lines)

def _inject_block(html, start_marker, end_marker, content):
    s = html.find(start_marker)
    e = html.find(end_marker)
    if s == -1 or e == -1:
        return html, False
    return html[:s + len(start_marker)] + '\n' + content + '\n' + html[e:], True

def inject_into_dashboard(data):
    try:
        html = open(DASHBOARD, 'r', encoding='utf-8').read()
    except FileNotFoundError:
        print('Dashboard not found: ' + DASHBOARD)
        return False

    ok_flags = []

    # 1. P2_DATA JS block
    js_block = 'const P2_DATA=' + json.dumps(data, indent=2) + ';'
    html, ok = _inject_block(html, '/* P2_DATA_START */', '/* P2_DATA_END */', js_block)
    ok_flags.append(('P2_DATA', ok))

    # 2. P0 Chevron context HTML block
    chev_ref  = load_chev_2025(data['day'])
    chev_html = build_chev_context_html(data['chev'], data['day'], data['date'], ref=chev_ref)
    html, ok = _inject_block(html, '<!-- P0_CHEV_START -->', '<!-- P0_CHEV_END -->', chev_html)
    ok_flags.append(('P0_CHEV', ok))

    # 3. Footers
    date_mixed, date_upper = _month_abbr(data['date'])
    day = data['day']

    p0_foot = ('<div class="p0-foot">TIMB FCV 2026 &nbsp;&middot;&nbsp; Season Day {}'
               ' &nbsp;&middot;&nbsp; 04 Mar &ndash; {}'
               ' &nbsp;&middot;&nbsp; STRICTLY INTERNAL</div>').format(day, date_mixed)
    html, ok = _inject_block(html, '<!-- P0_FOOT_START -->', '<!-- P0_FOOT_END -->', p0_foot)
    ok_flags.append(('P0_FOOT', ok))

    t_foot = ('<div class="t-foot">CHEVRON TOBACCO COMPANY (PVT) LTD &nbsp;&middot;&nbsp;'
              ' 2026 BUYING SEASON &nbsp;&nbsp;&middot;&nbsp;&nbsp; STRICTLY INTERNAL'
              ' &nbsp;&middot;&nbsp; SEASON DAY {} &nbsp;&middot;&nbsp; {}</div>').format(day, date_upper)
    html, ok = _inject_block(html, '<!-- T_FOOT_START -->', '<!-- T_FOOT_END -->', t_foot)
    ok_flags.append(('T_FOOT', ok))

    p2_foot = ('<div class="p2-foot">COMPETITOR INTELLIGENCE &nbsp;&middot;&nbsp; 2026 EXPORT MARKET'
               ' &nbsp;&middot;&nbsp; TIMB SECTOR ANALYSIS &nbsp;&middot;&nbsp; STRICTLY INTERNAL'
               ' &nbsp;&middot;&nbsp; SEASON DAY {}</div>').format(day)
    html, ok = _inject_block(html, '<!-- P2_FOOT_START -->', '<!-- P2_FOOT_END -->', p2_foot)
    ok_flags.append(('P2_FOOT', ok))

    for name, ok in ok_flags:
        if not ok:
            print('  WARNING: {} markers not found -- block skipped.'.format(name))

    # Atomic publish: write to <dest>.tmp then os.replace into the dashboard.
    # Mirrors the P1 pattern in generate_dashboard_data.py so both writers
    # behave the same way. P2 runs after a 15s orchestrator pause, so the
    # backoff here is shorter than P1's: 3 retries × 5s = ~15s max.
    # Note: P1 stamps the <meta name="build-version"> marker. P2 reads HTML
    # that already contains that marker and writes it back unchanged — no
    # marker handling needed here.
    import time as _time
    tmp_path = DASHBOARD + '.tmp'
    with open(tmp_path, 'w', encoding='utf-8') as f:
        f.write(html)
    for attempt in range(1, 4):          # 3 attempts ≈ 10s of backoff
        try:
            os.replace(tmp_path, DASHBOARD)
            break
        except PermissionError:
            if attempt == 3:
                # Clean up the orphan .tmp so OneDrive doesn't sync a sibling
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
                raise PermissionError(
                    "Dashboard file appears to be open or locked. "
                    "Close the browser/Excel preview and re-run."
                )
            print('  Dashboard locked (attempt {0}/3) - retrying in 5s...'.format(attempt))
            _time.sleep(5)
    return all(ok for _, ok in ok_flags)

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--day', type=int, default=None)
    parser.add_argument('--print-only', action='store_true')
    args = parser.parse_args()

    if args.day:
        path = os.path.join(DAILY_DIR, '2026 daily fcv summary day {}.xlsx'.format(args.day))
        day  = args.day
    else:
        path, day = find_latest_file()

    print('\nReading: {} (Day {})'.format(os.path.basename(path), day))
    data = extract(path, day)

    if args.print_only:
        print(json.dumps(data, indent=2))
    else:
        ok = inject_into_dashboard(data)
        status = 'OK - Dashboard updated' if ok else 'WARNING - Injection failed'
        print('\n{} -- Day {} - {}'.format(status, day, data['date']))
        print('Groups: {}'.format(len(data['groups'])))
        for g in data['groups']:
            print('  {:<20s}  {:>10,.0f} kg  ${:.4f}/kg  (con:{:>10,.0f}  auc:{:>10,.0f})'.format(
                g['name'], g['m'], g['p'], g['con_m'], g['auc_m']))
        print('\n  Chevron:              {:>10,.0f} kg  ${:.4f}/kg  (con:{:>10,.0f}  auc:{:>10,.0f})'.format(
            data['chev']['m'], data['chev']['p'], data['chev']['con_m'], data['chev']['auc_m']))
