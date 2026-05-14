"""
2026 TIMB FCV Market Data  Automated Extraction Script
--------------------------------------------------------
Scans the folder for all "2026*day*.xlsx" daily summary files,
extracts Auction, Contractor, and Sectoral data, and writes/overwrites
"Cleaned Tobacco Market Data 2026.xlsx" in the same folder.

Run this script manually or via a scheduled task whenever new daily
files arrive. It will detect the latest day and regenerate the output.

Optimisation notes (v2):
  - Persistent date-map cache (.cache_date_map.json)  only reads NEW daily files on each run
  - Single openpyxl read_only workbook load  avoids reading the large latest file 20+ times
  - Priority sheet order for date extraction ("Contract daily sales" first)
  - Typical run time after warm cache: ~20 seconds
"""

import os
import sys
import glob
import re
import json
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#  Configuration 
import os as _os
_WIN_ROOT   = r"C:\Users\MikeBingo\Chevron Leaf Tobacco\Power BI Uploads - 2026 TIMB FCV Daily Files"
_SCRIPT_DIR = _os.path.dirname(_os.path.abspath(__file__))
_LNX_ROOT   = _os.path.dirname(_SCRIPT_DIR)
ROOT_FOLDER = _WIN_ROOT if _os.path.exists(_WIN_ROOT) else _LNX_ROOT

FOLDER      = ROOT_FOLDER
OUTPUT_FILE = os.path.join(ROOT_FOLDER, "Cleaned Tobacco Market Data 2026.xlsx")
LOG_FILE    = os.path.join(ROOT_FOLDER, "market_data_refresh.log")
CACHE_FILE  = os.path.join(ROOT_FOLDER, ".cache_date_map.json")

AUCTION_FLOOR_SHEETS = {
    "TSF Daily Auction sales":  "TSF",
    "ETF Daily Auction sales":  "ETF",
    "PTSF Daily Auction sales": "PTSF",
}
CONTRACTOR_SHEETS = [
    "HARARE", "KAROI", "RUSAPE ", "MARONDERA",
    "MVURWI ", "MUTOKO", "BINDURA "
]
# Sheets that reliably contain "Date: DD/MM/YYYY" in the first 10 rows
DATE_PRIORITY_SHEETS = [
    "Contract daily sales", "Auction Daily Summary",
    "Seasonal Auction summary", "Seasonal Contract Sales",
]

#  Logging 
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


#  File helpers 
def discover_files(folder):
    pattern = os.path.join(folder, "2026*day*.xlsx")
    files = glob.glob(pattern)
    files = [f for f in files if re.search(r"day\s*\d+", f, re.IGNORECASE)]
    return sorted(files, key=lambda f: int(re.search(r"day\s*(\d+)", f, re.IGNORECASE).group(1)))


def day_num(path):
    return int(re.search(r"day\s*(\d+)", os.path.basename(path), re.IGNORECASE).group(1))


#  Date map (persistent cache) 
def _extract_date_from_file(path):
    """Extract the auction date from a daily summary file using pandas."""
    try:
        xl = pd.ExcelFile(path)
        sheet_order = [s for s in DATE_PRIORITY_SHEETS if s in xl.sheet_names]
        sheet_order += [s for s in xl.sheet_names if s not in DATE_PRIORITY_SHEETS]
        for sname in sheet_order:
            df = xl.parse(sname, nrows=10, header=None)
            flat = df.astype(str).values.flatten()
            for cell in flat:
                m = re.search(r"Date[:\s]+([\d]{1,2}[/\-][\d]{1,2}[/\-][\d]{2,4})", cell)
                if m:
                    try:
                        return str(pd.to_datetime(m.group(1), dayfirst=True).date())
                    except Exception:
                        pass
    except Exception as e:
        log.warning(f"Date extraction failed for {os.path.basename(path)}: {e}")
    return None


def load_date_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE) as f:
                return {int(k): v for k, v in json.load(f).items()}
        except Exception:
            pass
    return {}


def save_date_cache(dm):
    with open(CACHE_FILE, "w") as f:
        json.dump({str(k): v for k, v in dm.items()}, f)


def build_date_map(files):
    """Load cached dates; only read files not yet in the cache."""
    cache = load_date_cache()
    missing = [f for f in files if day_num(f) not in cache]
    if missing:
        log.info(f"  Fetching dates for {len(missing)} new file(s)...")
        for path in missing:
            d = day_num(path)
            cache[d] = _extract_date_from_file(path)
            log.info(f"    Day {d:>3} = {cache[d]}")
        save_date_cache(cache)
    else:
        log.info(f"  All {len(cache)} dates loaded from cache.")
    return {k: pd.Timestamp(v) if v else pd.NaT for k, v in cache.items()}


#  Sheet extraction (from pre-loaded openpyxl workbook) 
def _wb_sheet_to_df(wb, sheet_name, skip_rows=0, nrows=None):
    """Read a worksheet into a DataFrame via openpyxl (no repeated file I/O)."""
    ws = wb[sheet_name]
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_rows:
            continue
        if nrows is not None and (i - skip_rows) >= nrows:
            break
        data.append(list(row))
    return pd.DataFrame(data) if data else pd.DataFrame()


def extract_wide_sheet(wb, sheet_name, max_day):
    """Read a wide-format sheet (days as column groups) and return long-format DataFrame."""
    hdr = _wb_sheet_to_df(wb, sheet_name, skip_rows=0, nrows=4)
    day_row    = pd.Series(hdr.iloc[2]).ffill()
    metric_row = pd.Series(hdr.iloc[3])
    mi = pd.MultiIndex.from_arrays([day_row, metric_row])

    df_raw = _wb_sheet_to_df(wb, sheet_name, skip_rows=4)
    if df_raw.empty:
        raise ValueError(f"No data rows in sheet '{sheet_name}'")

    if len(mi) < df_raw.shape[1]:
        extra = [("Unknown", f"Extra_{i+1}") for i in range(df_raw.shape[1] - len(mi))]
        mi = mi.append(pd.MultiIndex.from_tuples(extra))
    elif len(mi) > df_raw.shape[1]:
        mi = mi[:df_raw.shape[1]]

    df_raw.columns = mi
    df_flat = df_raw.copy()
    df_flat.columns = [f"{a}||{b}" for a, b in df_flat.columns]

    company_col = next((c for c in df_flat.columns if "company" in str(c).lower()), None)
    if not company_col:
        raise ValueError("No 'Company' column found.")

    df_melted = df_flat.melt(id_vars=[company_col], var_name="Day_Metric", value_name="Value")
    df_melted.rename(columns={company_col: "Company"}, inplace=True)
    df_melted[["Day", "Metric"]] = df_melted["Day_Metric"].str.extract(r"Day\s*(\d+).*?\|\|(.*)")
    df_melted.drop(columns=["Day_Metric"], inplace=True)
    df_melted["Day"] = pd.to_numeric(df_melted["Day"], errors="coerce").astype("Int64")
    df_melted = df_melted[df_melted["Day"] <= max_day]

    df_tidy = df_melted.pivot_table(
        index=["Company", "Day"], columns="Metric", values="Value", aggfunc="first"
    ).reset_index()
    df_tidy.columns.name = None
    return df_tidy


#  Extraction functions 
def extract_auction(wb, date_map):
    max_day = max(date_map.keys())
    frames = []

    for sheet_name, floor_label in AUCTION_FLOOR_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            log.warning(f"Auction sheet '{sheet_name}' not found  skipping.")
            continue
        try:
            df = extract_wide_sheet(wb, sheet_name, max_day)

            for col in ["Mass (kg)", "Value (US$)", "Bales Sold", "% Mass", "Ave. BaleWeight"]:
                if col not in df.columns:
                    df[col] = pd.NA
                df[col] = pd.to_numeric(df[col], errors="coerce")

            df["Ave. (US$/kg)"]  = df["Value (US$)"] / df["Mass (kg)"]
            df["Ave. BaleWeight"] = df["Mass (kg)"]  / df["Bales Sold"]
            day_totals = df.groupby("Day")["Mass (kg)"].transform("sum")
            df["% Mass"] = (df["Mass (kg)"] / day_totals) * 100

            df = df[(df["Mass (kg)"] > 0) | (df["Value (US$)"] > 0)].copy()
            df = df[~df["Company"].astype(str).str.upper().str.contains("TOTAL|ALL FLOORS", na=False)]

            df.insert(df.columns.get_loc("Company"), "Auction Floor", floor_label)
            df["Date of Sale"] = df["Day"].map(date_map)

            frames.append(df)
            log.info(f"  Auction {floor_label:<6}: {len(df):>4} rows")
        except Exception as e:
            log.error(f"  Auction '{sheet_name}' failed: {e}")

    combined = pd.concat(frames, ignore_index=True)
    cols = ["Auction Floor", "Company", "Day", "% Mass", "Ave. (US$/kg)",
            "Ave. BaleWeight", "Bales Sold", "Mass (kg)", "Value (US$)", "Date of Sale"]
    return combined[[c for c in cols if c in combined.columns]]


def extract_contractor(wb, date_map):
    max_day = max(date_map.keys())
    frames = []

    for sheet_name in CONTRACTOR_SHEETS:
        if sheet_name not in wb.sheetnames:
            log.warning(f"Contractor sheet '{sheet_name}' not found  skipping.")
            continue
        try:
            df = extract_wide_sheet(wb, sheet_name, max_day)

            for col in ["Mass (kg)", "Value (US$)", "Bales Sold"]:
                if col not in df.columns:
                    df[col] = pd.NA
                df[col] = pd.to_numeric(df[col], errors="coerce")

            df["Ave. (US$/kg)"] = (df["Value (US$)"] / df["Mass (kg)"]).round(2)

            df = df[(df["Mass (kg)"] > 0) | (df["Value (US$)"] > 0)].copy()
            df = df[~df["Company"].astype(str).str.upper().str.contains("TOTAL", na=False)]

            df.insert(df.columns.get_loc("Company"), "Region", sheet_name.strip().upper())
            df["Date of Sale"] = df["Day"].map(date_map)

            frames.append(df)
            log.info(f"  Contractor {sheet_name.strip():<12}: {len(df):>4} rows")
        except Exception as e:
            log.error(f"  Contractor '{sheet_name}' failed: {e}")

    combined = pd.concat(frames, ignore_index=True)
    cols = ["Region", "Company", "Day", "Date of Sale", "Mass (kg)", "Value (US$)", "Bales Sold", "Ave. (US$/kg)"]
    return combined[[c for c in cols if c in combined.columns]]


def extract_sectoral(wb):
    raw = _wb_sheet_to_df(wb, "Seasonal Sales By Sector", skip_rows=2)
    raw.columns = range(raw.shape[1])
    raw = raw[raw[0].notna()]
    raw = raw[~raw[0].astype(str).str.upper().str.contains("TOTAL", na=False)]
    raw = raw[~raw[0].astype(str).str.contains("MASS|VALUE|Avg", case=False, na=False)]
    raw.rename(columns={0: "Company"}, inplace=True)

    small = raw[["Company", 1, 2, 3]].copy()
    small.columns = ["Company", "Mass (kg)", "Value (US$)", "Ave. Price (US$/kg)"]
    small["Sector"] = "Small Scale"

    comm = raw[["Company", 4, 5, 6]].copy()
    comm.columns = ["Company", "Mass (kg)", "Value (US$)", "Ave. Price (US$/kg)"]
    comm["Sector"] = "Commercial"

    combined = pd.concat([small, comm], ignore_index=True)
    for col in ["Mass (kg)", "Value (US$)", "Ave. Price (US$/kg)"]:
        combined[col] = pd.to_numeric(combined[col], errors="coerce")

    result = combined[(combined["Mass (kg)"] > 0) | (combined["Value (US$)"] > 0)].reset_index(drop=True)
    log.info(f"  Sectoral: {len(result)} rows  ({result['Sector'].value_counts().to_dict()})")
    return result


#  Excel output 
def write_excel(auction_df, contractor_df, sectoral_df, output_path):
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HDR_FILL  = PatternFill("solid", start_color="1F4E79")
    DATA_FONT = Font(name="Arial", size=10)
    ALT_FILL  = PatternFill("solid", start_color="DCE6F1")
    TOT_FONT  = Font(name="Arial", bold=True, size=10)
    TOT_FILL  = PatternFill("solid", start_color="BDD7EE")
    thin      = Side(style="thin", color="B8CCE4")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")

    def write_sheet(ws, df, num_fmts, col_widths, date_cols=None):
        headers = list(df.columns)
        n = len(headers)
        date_cols = date_cols or []

        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[1].height = 20

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            alt = (r_idx % 2 == 0)
            for c_idx, col in enumerate(headers, 1):
                val = row[col]
                cell = ws.cell(row=r_idx, column=c_idx)
                if col in date_cols:
                    cell.value = pd.Timestamp(val).to_pydatetime() if pd.notna(val) else None
                    cell.number_format = "DD/MM/YYYY"
                else:
                    cell.value = None if (not isinstance(val, str) and pd.isna(val)) else val
                cell.font   = DATA_FONT
                cell.border = BORDER
                cell.fill   = ALT_FILL if alt else PatternFill()
                if col in num_fmts:
                    cell.number_format = num_fmts[col]
                cell.alignment = LEFT if (isinstance(val, str) or col in date_cols) else RIGHT

        tr = len(df) + 2
        ws.cell(row=tr, column=1, value="TOTAL").font = TOT_FONT
        ws.cell(row=tr, column=1).fill = TOT_FILL
        ws.cell(row=tr, column=1).border = BORDER
        ws.cell(row=tr, column=1).alignment = LEFT
        for c_idx, col in enumerate(headers[1:], 2):
            cell = ws.cell(row=tr, column=c_idx)
            cell.font = TOT_FONT; cell.fill = TOT_FILL; cell.border = BORDER
            if df[col].dtype in ["float64", "int64"] and col not in date_cols:
                cl = get_column_letter(c_idx)
                cell.value = f"=SUM({cl}2:{cl}{tr-1})"
                cell.alignment = RIGHT
                if col in num_fmts:
                    cell.number_format = num_fmts[col]
            else:
                cell.value = ""; cell.alignment = LEFT

        for col, width in col_widths.items():
            if col in headers:
                ws.column_dimensions[get_column_letter(headers.index(col)+1)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(n)}1"

    wb_out = Workbook()

    ws1 = wb_out.active
    ws1.title = "Auction Data 2026"
    write_sheet(ws1, auction_df,
        num_fmts={"% Mass": "0.000000", "Ave. (US$/kg)": "#,##0.00",
                  "Ave. BaleWeight": "#,##0.00", "Bales Sold": "#,##0",
                  "Mass (kg)": "#,##0", "Value (US$)": "#,##0.00"},
        col_widths={"Auction Floor": 14, "Company": 32, "Day": 7, "% Mass": 12,
                    "Ave. (US$/kg)": 14, "Ave. BaleWeight": 16, "Bales Sold": 12,
                    "Mass (kg)": 14, "Value (US$)": 16, "Date of Sale": 14},
        date_cols=["Date of Sale"])

    ws2 = wb_out.create_sheet("Contractor Data 2026")
    write_sheet(ws2, contractor_df,
        num_fmts={"Mass (kg)": "#,##0", "Value (US$)": "#,##0.00",
                  "Bales Sold": "#,##0", "Ave. (US$/kg)": "#,##0.00"},
        col_widths={"Region": 14, "Company": 36, "Day": 7, "Date of Sale": 14,
                    "Mass (kg)": 16, "Value (US$)": 16, "Bales Sold": 12, "Ave. (US$/kg)": 14},
        date_cols=["Date of Sale"])

    ws3 = wb_out.create_sheet("Sectoral Data 2026")
    write_sheet(ws3, sectoral_df,
        num_fmts={"Mass (kg)": "#,##0.00", "Value (US$)": "#,##0.00",
                  "Ave. Price (US$/kg)": "#,##0.00"},
        col_widths={"Company": 40, "Mass (kg)": 18, "Value (US$)": 18,
                    "Ave. Price (US$/kg)": 20, "Sector": 14})

    wb_out.save(output_path)


#  Dashboard updates
def _run_dashboard_updates(latest_file, latest_day, root_folder, log):
    """Run both dashboard injectors via subprocess — P1 first (5 min timeout), pause, then P2."""
    import subprocess as _sp
    import time as _time

    _py  = sys.executable
    _env = os.environ.copy()
    _env["PYTHONUTF8"] = "1"   # force UTF-8 console on Windows — fixes charmap errors

    P1_TIMEOUT = 300   # 5 minutes
    P2_TIMEOUT = 120   # 2 minutes

    #  P1: Regional & Overview (writes HTML first)
    p1_script = os.path.join(root_folder, "Scripts", "generate_dashboard_data.py")
    log.info(f"Running P1 (generate_dashboard_data.py)...")
    try:
        r1 = _sp.run(
            [_py, p1_script],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=P1_TIMEOUT, env=_env
        )
        if r1.returncode == 0:
            log.info(f"P1 updated -- Day {latest_day}")
        else:
            log.warning(f"P1 exited with code {r1.returncode}")
            if r1.stderr.strip():
                log.warning(f"P1 stderr: {r1.stderr.strip()[-500:]}")
    except _sp.TimeoutExpired:
        log.warning(f"P1 timed out after {P1_TIMEOUT}s — skipping")
    except Exception as e:
        log.warning(f"P1 failed: {e}")

    #  Pause to let OneDrive release the HTML before P2 writes to it
    log.info("Pausing 15s for OneDrive to release HTML...")
    _time.sleep(15)

    #  P2: Competitor Intelligence (writes HTML second)
    p2_script = os.path.join(root_folder, "Scripts", "generate_competitor_data.py")
    log.info(f"Running P2 (generate_competitor_data.py)...")
    try:
        r2 = _sp.run(
            [_py, p2_script],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=P2_TIMEOUT, env=_env
        )
        if r2.returncode == 0:
            log.info(f"P2 updated -- Day {latest_day}")
        else:
            log.warning(f"P2 exited with code {r2.returncode}")
            if r2.stderr.strip():
                log.warning(f"P2 stderr: {r2.stderr.strip()[-500:]}")
    except _sp.TimeoutExpired:
        log.warning(f"P2 timed out after {P2_TIMEOUT}s — skipping")
    except Exception as e:
        log.warning(f"P2 failed: {e}")


#  Git push 
def _git_push(root_folder, day, log):
    import subprocess, shutil
    dashboard = os.path.join(root_folder, "Market Intelligence Dashboard 2026.html")
    index     = os.path.join(root_folder, "index.html")
    try:
        shutil.copy2(dashboard, index)
        log.info("Copied dashboard -> index.html")
    except Exception as e:
        log.warning(f"Could not copy to index.html: {e}")
        return
    def _git(args):
        return subprocess.run(["git", "-C", root_folder] + args, capture_output=True, text=True)
    _git(["add", "Market Intelligence Dashboard 2026.html", "index.html"])
    commit = _git(["commit", "-m", f"Dashboard auto-update -- Day {day}"])
    if "nothing to commit" in commit.stdout + commit.stderr:
        log.info("Git: nothing new to commit.")
        return
    push = _git(["push"])
    if push.returncode == 0:
        log.info(f"GitHub Pages pushed -- Day {day}")
    else:
        log.warning(f"Git push failed: {push.stderr.strip()}")


def main():
    import time
    t0 = time.time()
    log.info("=" * 60)
    log.info("2026 TIMB FCV Market Data Refresh -- START")
    log.info(f"Folder : {FOLDER}")
    files = discover_files(FOLDER)
    if not files:
        log.error("No 2026 daily FCV files found. Exiting.")
        sys.exit(1)
    log.info(f"Files found: {len(files)}  (Day 1 -> Day {len(files)})")
    latest_file = files[-1]
    latest_day  = day_num(latest_file)
    if os.path.exists(OUTPUT_FILE):
        out_mtime    = os.path.getmtime(OUTPUT_FILE)
        latest_mtime = os.path.getmtime(latest_file)
        if latest_mtime <= out_mtime:
            log.info(f"Output already up-to-date (last day: {latest_day}). Skipping extraction.")
            log.info("  Running dashboard updates anyway...")
            _run_dashboard_updates(latest_file, latest_day, ROOT_FOLDER, log)
            _git_push(ROOT_FOLDER, latest_day, log)
            log.info("=" * 60)
            return
        log.info(f"New data detected (Day {latest_day}). Regenerating...")
    else:
        log.info(f"Output file does not exist. Building from {len(files)} files...")
    log.info("Building date map (cached)...")
    date_map = build_date_map(files)
    log.info(f"  Date map: Day {min(date_map)} -> Day {max(date_map)}  [{time.time()-t0:.1f}s]")
    log.info(f"Loading workbook (read_only)... [{time.time()-t0:.1f}s]")
    wb = load_workbook(latest_file, read_only=True, data_only=True)
    log.info(f"  Workbook loaded. Sheets: {len(wb.sheetnames)}  [{time.time()-t0:.1f}s]")
    log.info(f"Extracting Auction data...  [{time.time()-t0:.1f}s]")
    auction_df = extract_auction(wb, date_map)
    log.info(f"Extracting Contractor data...  [{time.time()-t0:.1f}s]")
    contractor_df = extract_contract